"""
import_excel.py
----------------
Loads legacy eBill data (an Excel export OR a direct Access .mdb/.accdb file)
into the web app's database. Both formats end up shaped the same way -- a
set of named tables, each with a header row and data rows -- so a single
mapping function (_run_from_source) drives both. Only the "reader" differs:

  - Excel (.xlsx): openpyxl reads each worksheet.
  - Access (.mdb/.accdb): tried two ways, in order --
      1. pyodbc + the "Microsoft Access Driver" -- this is the path that
         works on WINDOWS (where most people actually have the .mdb file
         and no apt-get). Needs `pip install pyodbc` plus the free
         "Microsoft Access Database Engine Redistributable" installed
         (32-bit or 64-bit, matching your Python install).
      2. mdbtools' `mdb-tables` / `mdb-export` command-line utilities --
         this is the path for LINUX/MAC/WSL servers
         (apt-get install mdbtools / brew install mdbtools).
    If neither is available, run_mdb_import() raises MdbAccessUnavailable
    with a message explaining both options plus the Excel-export fallback.

Historical vouchers (Production Vouchers, Delivery Challans, Tax Invoices,
Purchase Bills, Battery Delivery Challans, Old Rickshaw buy-backs, Day Book
entries) ARE now imported, alongside master data (Dealers, Products, simple
masters, Production Formula/BOM, the Chassis/Vehicle register, and Users)
plus the Company profile.

Day Book / the legacy general-ledger "FT" table: the exact column names of
that table in your real Access/Excel export haven't been confirmed against
a sample file, so the matcher below tries several plausible aliases (VRNO,
DATE, PARTY/DNAME, CR/CRAMT, DR/DRAMT, NARRATION/REMARK, etc.). If your
Day Book count comes back 0 after an import that clearly has that data,
send a sample of that sheet/table's column headers and this can be
tightened up.
"""
import csv
import io
import subprocess
from collections import defaultdict
from datetime import datetime as _dt

import openpyxl
from models import (
    db, Dealer, Product, SimpleMaster, Vehicle, User, Company, ProductionFormula,
    ProductionVoucher, DeliveryChallan, TaxInvoice, PurchaseBill, PurchaseBillItem,
    OldRickshaw, BatteryDeliveryChallan, DayBook,
)

try:
    import pyodbc
except ImportError:
    pyodbc = None


class MdbToolsMissing(Exception):
    """Raised when the mdbtools command-line utilities aren't installed."""
    pass


class AccessDriverMissing(Exception):
    """Raised when pyodbc or the Windows Access ODBC driver isn't available."""
    pass


class MdbAccessUnavailable(Exception):
    """Raised when NEITHER the pyodbc/Windows path NOR mdbtools worked."""
    pass


# ---------------------------------------------------------------------------
# Small coercion helpers -- Excel gives us native datetime/float objects,
# mdb-export gives us CSV strings, so both readers funnel through these.
# ---------------------------------------------------------------------------
def _to_date(value):
    if value is None or value == "":
        return None
    if hasattr(value, "date") and callable(getattr(value, "date")):
        return value.date()
    if hasattr(value, "year"):  # already a date
        return value
    for fmt in ("%m/%d/%y", "%m/%d/%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return _dt.strptime(str(value).strip(), fmt).date()
        except ValueError:
            continue
    return None


def _to_float(value):
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _s(value):
    """Trim a legacy fixed-width CHAR field; blank/None -> None."""
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _b(value):
    """Legacy Y/N-ish flag -> bool. Blank, None, or ' ' all mean False."""
    return str(value).strip().upper() == "Y" if value is not None else False


def _batno(value):
    """Battery serial columns use 0 (or blank) to mean 'not assigned'."""
    if value in (None, 0, "0", ""):
        return None
    return str(value).strip()


_RELATION_PREFIXES = ("S/o", "D/o", "W/o", "C/o")


def _split_relation(value):
    """
    Legacy SFATHER field packs relation + name into one string, e.g.
    'W/o AVDESH KUMAR' or 'S/o VINOD KUMAR BANSAL' -- split it into
    TaxInvoice's separate buyer_relation / buyer_father_name columns.
    Falls back to the "S/o" model default if there's no recognised
    prefix (blank, or a name with no relation tag at all).
    """
    s = _s(value)
    if not s:
        return "S/o", None
    parts = s.split(None, 1)
    if len(parts) == 2 and parts[0] in _RELATION_PREFIXES:
        return parts[0], parts[1].strip() or None
    return "S/o", s


# ---------------------------------------------------------------------------
# Table sources -- both expose find_table(required_columns) -> (header, rows)
# where rows is a list of dicts already keyed by column name. This lets the
# mapping logic below stay 100% identical for Excel and Access.
# ---------------------------------------------------------------------------
class ExcelSource:
    def __init__(self, filepath):
        self.wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    def find_table(self, required_columns):
        """
        Find the sheet whose header row contains all required_columns. Some
        legacy sheets are wide, denormalized voucher tables that happen to
        ALSO contain master-data columns. To avoid matching those by
        mistake, if more than one sheet qualifies we pick the one with the
        FEWEST columns overall -- the real master table is always the
        narrow one.
        """
        candidates = []
        for sn in self.wb.sheetnames:
            ws = self.wb[sn]
            try:
                first_row = next(ws.iter_rows(min_row=1, max_row=1))
            except StopIteration:
                continue  # empty sheet
            header = [c.value for c in first_row]
            if all(col in header for col in required_columns):
                candidates.append((ws, header))
        if not candidates:
            return None, None
        candidates.sort(key=lambda pair: len(pair[1]))
        ws, header = candidates[0]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None or all(v is None for v in row):
                continue
            rows.append(dict(zip(header, row)))
        return header, rows


class MdbSource:
    """Reads tables out of an Access .mdb/.accdb file via the mdbtools CLI."""

    def __init__(self, filepath, password=None):
        # mdbtools' CLI utilities have no built-in support for
        # password-protected Access databases -- there is no flag to
        # supply one. If the caller told us there's a password, don't
        # even try; fail fast with a clear explanation instead of a
        # confusing low-level mdbtools error.
        if password:
            raise MdbToolsMissing(
                "this file is password-protected, and mdbtools has no way "
                "to supply a password -- use the Windows/pyodbc route instead"
            )
        self.filepath = filepath
        try:
            out = subprocess.run(["mdb-tables", "-1", filepath],
                                  capture_output=True, text=True, check=True)
        except FileNotFoundError:
            raise MdbToolsMissing()
        except subprocess.CalledProcessError as e:
            raise RuntimeError(f"mdb-tables failed: {e.stderr}")
        self.table_names = [t for t in out.stdout.splitlines() if t.strip()]
        self._cache = {}

    def _read_table(self, name):
        if name in self._cache:
            return self._cache[name]
        try:
            out = subprocess.run(["mdb-export", self.filepath, name],
                                  capture_output=True, text=True, check=True)
        except FileNotFoundError:
            raise MdbToolsMissing()
        except subprocess.CalledProcessError as e:
            raise RuntimeError(f"mdb-export failed for table '{name}': {e.stderr}")
        rows_raw = list(csv.reader(io.StringIO(out.stdout)))
        if not rows_raw:
            self._cache[name] = (None, [])
            return None, []
        header = rows_raw[0]
        rows = [dict(zip(header, r)) for r in rows_raw[1:]]
        self._cache[name] = (header, rows)
        return header, rows

    def find_table(self, required_columns):
        candidates = []
        for name in self.table_names:
            header, rows = self._read_table(name)
            if header and all(col in header for col in required_columns):
                candidates.append((header, rows))
        if not candidates:
            return None, None
        candidates.sort(key=lambda pair: len(pair[0]))
        return candidates[0]


class AccessOdbcSource:
    """
    Reads tables out of an Access .mdb/.accdb file using pyodbc + the
    "Microsoft Access Driver". This is the path that works on WINDOWS,
    where mdbtools generally isn't installed. Requires:
      - `pip install pyodbc`
      - the free "Microsoft Access Database Engine Redistributable"
        installed (pick the 32-bit or 64-bit build to match your Python
        install -- mismatched bitness is the #1 cause of driver errors)
    """

    def __init__(self, filepath, password=None):
        if pyodbc is None:
            raise AccessDriverMissing("pyodbc is not installed (pip install pyodbc).")
        # NOTE: unlike some ODBC drivers, the Microsoft Access driver takes
        # PWD literally -- it does NOT support {}-escaping for special
        # characters. Wrapping the password in braces here would make the
        # driver treat the braces as part of the password and fail with a
        # misleading "Not a valid password" error even for the right one.
        # If a password legitimately contains a literal ";" this simple
        # approach can't express it -- that's a known limitation.
        pwd = password or ""
        conn_str = (
            r"DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};"
            rf"DBQ={filepath};"
            rf"PWD={pwd};"
        )
        try:
            self.conn = pyodbc.connect(conn_str, autocommit=True)
        except pyodbc.Error as e:
            hint = (
                " The database appears to be password-protected -- enter the "
                "password and try again."
                if not password else
                " Double-check the password is correct, or (if it's blank) that "
                "the file doesn't actually require one."
            )
            raise AccessDriverMissing(
                "Could not open the file via the Microsoft Access ODBC driver "
                f"(is it installed, and does its bitness match your Python?).{hint} Details: {e}"
            )
        self._cache = {}

    def _table_names(self):
        cursor = self.conn.cursor()
        # Skip Access system tables (MSys...) -- they're never what we want.
        return [row.table_name for row in cursor.tables(tableType="TABLE")
                if not row.table_name.startswith("MSys")]

    def _read_table(self, name):
        if name in self._cache:
            return self._cache[name]
        cursor = self.conn.cursor()
        try:
            cursor.execute(f"SELECT * FROM [{name}]")
        except pyodbc.Error:
            self._cache[name] = (None, [])
            return None, []
        header = [col[0] for col in cursor.description]
        rows = [dict(zip(header, row)) for row in cursor.fetchall()]
        self._cache[name] = (header, rows)
        return header, rows

    def find_table(self, required_columns):
        candidates = []
        for name in self._table_names():
            header, rows = self._read_table(name)
            if header and all(col in header for col in required_columns):
                candidates.append((header, rows))
        if not candidates:
            return None, None
        candidates.sort(key=lambda pair: len(pair[0]))
        return candidates[0]


# ---------------------------------------------------------------------------
# Shared mapping logic -- identical regardless of source type.
# ---------------------------------------------------------------------------
def _run_from_source(source):
    summary = {}

    # --- Company profile (Sheet/table with NAME/GSTNO/BANKNAME...) --------
    header, rows = source.find_table(["NAME", "GSTNO", "BANKNAME"])
    if header:
        for r in rows:
            c = Company.query.first() or Company()
            c.name = r.get("NAME") or c.name
            c.year = r.get("YEAR") or c.year
            c.address1 = r.get("ADD1")
            c.address2 = r.get("ADD2")
            c.state = r.get("STATE")
            c.mobile = r.get("MOB")
            c.website = r.get("WEBSITE")
            c.gst_no = r.get("GSTNO")
            c.bank_name = r.get("BANKNAME")
            c.bank_account = r.get("BANKAC")
            c.bank_ifsc = r.get("BANKIFSC")
            c.state_code = r.get("STCODE")
            db.session.add(c)
            break  # single-row table
        db.session.commit()

    # --- Dealers (AMC/AMN/...) ---------------------------------------------
    header, rows = source.find_table(["AMC", "AMN", "AMCODE"])
    count = 0
    if header:
        for r in rows:
            code = str(r.get("AMCODE") or "").strip() or None
            existing = Dealer.query.filter_by(code=code).first() if code else None
            d = existing or Dealer()
            d.code = code
            d.name = r.get("AMN") or ""
            d.address1 = r.get("ADD1")
            d.address2 = r.get("ADD2")
            d.mobile = r.get("MOB")
            d.gst_no = r.get("GSTNO")
            d.state = r.get("STATE")
            d.state_code = r.get("STCODE")
            d.pan = r.get("PAN")
            d.salesman = _s(r.get("SALESMAN"))
            if d.name:
                db.session.add(d)
                count += 1
        db.session.commit()
    summary["Dealers"] = count

    # --- Products / Items (IMC/IMN/UNIT/GST/HSN) ---------------------------
    header, rows = source.find_table(["IMC", "IMN", "UNIT"])
    count = 0
    if header:
        for r in rows:
            code = str(r.get("IMC") or "").strip() or None
            existing = Product.query.filter_by(code=code).first() if code else None
            p = existing or Product()
            p.code = code
            p.name = r.get("IMN") or ""
            p.unit = r.get("UNIT") or "PCS"
            p.gst_rate = _to_float(r.get("GST"))
            p.hsn_code = r.get("HSN")
            p.fro = r.get("FRO") or "R"
            # Legacy column name for this isn't confirmed from a live sample yet --
            # try the likely variants. If none hit, umrn_code will just stay blank
            # after import; check the actual source table's column name and adjust
            # this line if so.
            p.umrn_code = _s(r.get("UMRN")) or _s(r.get("UMRNCODE")) or _s(r.get("UMRNNO"))
            if p.name:
                db.session.add(p)
                count += 1
        db.session.commit()
    summary["Products"] = count

    # --- Simple masters: Colour / Battery Maker / Financer / Mechanic / ----
    # --- RTO / Party (Raw Material Purchase Parties) ------------------------
    address_col_by_kind = {"financer": "HADD1", "rto": "CADD1", "party": "ADD1"}
    mobile_col_by_kind = {"party": "MOB"}
    extra_col_by_kind = {"party": "GSTNO"}
    simple_specs = [
        ("colour", ["COLOUR", "COLOUR_CODE"], "COLOUR", "COLOUR_CODE"),
        ("battery-maker", ["BMC", "BMN"], "BMN", "BMC"),
        ("financer", ["HMC", "HMN"], "HMN", "HMC"),
        ("mechanic", ["MMC", "MMN"], "MMN", "MMC"),
        ("rto", ["CMC", "CMN"], "CMN", "CMC"),
        ("party", ["VMC", "VMN"], "VMN", "VMC"),
    ]
    for kind, required, name_col, code_col in simple_specs:
        header, rows = source.find_table(required)
        count = 0
        if header:
            for r in rows:
                name = r.get(name_col)
                if not name:
                    continue
                code = str(r.get(code_col) or "")
                existing = SimpleMaster.query.filter_by(kind=kind, code=code).first() if code else None
                row = existing or SimpleMaster(kind=kind)
                row.name = str(name)
                row.code = code
                row.address = r.get(address_col_by_kind[kind]) if kind in address_col_by_kind else None
                row.mobile = r.get(mobile_col_by_kind[kind]) if kind in mobile_col_by_kind else None
                row.extra = r.get(extra_col_by_kind[kind]) if kind in extra_col_by_kind else None
                db.session.add(row)
                count += 1
            db.session.commit()
        summary[f"SimpleMaster:{kind}"] = count

    # --- Production Formula / BOM (PFC/PFN/FIMC/RIMC/WT) -------------------
    # NOTE: the legacy PFM table has NO raw-item-name column (no "RIMN") --
    # only codes (FIMC = finished item code, RIMC = raw item code). The raw
    # item's display name has to come from the Product master via RIMC.
    header, rows = source.find_table(["PFC", "FIMC", "RIMC", "WT"])
    count = 0
    if header:
        product_by_code_local = {p.code: p for p in Product.query.all() if p.code}
        ProductionFormula.query.delete()  # rebuild cleanly each import
        for r in rows:
            fimc = str(r.get("FIMC") or "")
            rimc = str(r.get("RIMC") or "")
            finished = product_by_code_local.get(fimc)
            raw_item = product_by_code_local.get(rimc)
            product_name = r.get("PFN") or (finished.name if finished else None)
            if not product_name:
                continue
            row = ProductionFormula(
                # Legacy data has no separate formula name -- it's the same
                # as the finished product name, same as the legacy app.
                formula_name=str(product_name),
                product_code=fimc,
                product_name=str(product_name),
                raw_item_code=rimc,
                raw_item_name=raw_item.name if raw_item else rimc,
                qty=_to_float(r.get("WT")) or 1,
                unit=(raw_item.unit if raw_item else None) or "PCS",
            )
            db.session.add(row)
            count += 1
        db.session.commit()
    summary["ProductionFormula lines"] = count

    # --- Vehicles / Chassis register (DT/IMC/IMN/CHASSIS/MOTOR/...) --------
    # NOTE: legacy fixed-width fields (Access CHAR columns) come back
    # space-padded -- e.g. CHASSIS as 'MD9GRDDE2LD245815   '. Every other
    # import block strips identifier fields with _s() before using them as
    # a lookup key, so this table has to as well: if chassis_no were stored
    # padded here while lookups elsewhere strip it first, every vehicle
    # lookup downstream (Delivery Challans, Tax Invoices, Production
    # Vouchers) would silently fail to match -- no error, just vehicles
    # that never link up.
    header, rows = source.find_table(["CHASSIS", "MOTOR", "CONTROLLER", "COLOUR"])
    count = 0
    if header:
        for r in rows:
            chassis_no = _s(r.get("CHASSIS"))
            if not chassis_no:
                continue
            existing = Vehicle.query.filter_by(chassis_no=chassis_no).first()
            v = existing or Vehicle()
            v.chassis_no = chassis_no
            # Every row for a given chassis (Production, Delivery, Sale,
            # Battery...) shares this same source table, but only the
            # Production-stage row usually carries model/motor/colour/etc.
            # -- later rows for the same chassis are typically blank on
            # these fields. Keep the previously-set value instead of
            # letting a later blank row wipe out a correct earlier one.
            if not v.date:
                v.date = _to_date(r.get("DT"))
            v.model_name = _s(r.get("IMN")) or v.model_name
            v.motor_no = _s(r.get("MOTOR")) or v.motor_no
            v.controller_no = _s(r.get("CONTROLLER")) or v.controller_no
            v.differential_no = _s(r.get("DIFFERENTIAL")) or v.differential_no
            v.colour = _s(r.get("COLOUR")) or v.colour
            v.colour_code = r.get("COLOUR_CODE") or v.colour_code
            v.other = _s(r.get("OTHNO")) or v.other
            v.stage = v.stage or "Manufacturing"
            db.session.add(v)
            count += 1
        db.session.commit()
    summary["Vehicles"] = count

    # -------------------------------------------------------------------
    # Historical vouchers.
    #
    # In the legacy database almost all vouchers (Production, Delivery
    # Challan, Tax Invoice, Purchase Bill) live together in ONE wide table
    # (Access table name "INV" -- the same table the Vehicles/Chassis
    # register above was built from), distinguished by a "VR" column:
    #   VR='W' -> Production Voucher     VR='S' -> Tax Invoice
    #   VR='D' -> Delivery Challan       VR='P' -> Purchase Bill
    #   VR='B' -> a battery-replacement voucher, handled together with
    #             the small dedicated Battery Delivery Challan table below.
    # Old Rickshaw (buy-back) records live in their own table (VEHNO/
    # ONAME/SOLDAMT/LOANAMT), separate from the above.
    # -------------------------------------------------------------------

    # Lookup tables built from already-imported masters, keyed the way the
    # voucher rows reference them (numeric internal codes, not display codes).
    dealer_by_code = {d.code: d for d in Dealer.query.all() if d.code}
    product_by_code = {p.code: p for p in Product.query.all() if p.code}
    battery_by_code = {s.code: s for s in SimpleMaster.query.filter_by(kind="battery-maker") if s.code}
    financer_by_code = {s.code: s for s in SimpleMaster.query.filter_by(kind="financer") if s.code}
    colour_code_by_name = {s.name.strip(): s.code for s in SimpleMaster.query.filter_by(kind="colour") if s.name}

    # AM/VM/CM reference numeric-id -> display fields that aren't stored on
    # our Dealer/SimpleMaster rows themselves, so read them fresh from source.
    amc_header, amc_rows = source.find_table(["AMC", "AMN", "AMCODE"])
    dealer_by_amc = {}
    if amc_header:
        for r in amc_rows:
            amcode = str(r.get("AMCODE") or "").strip()
            if r.get("AMC") is not None and amcode in dealer_by_code:
                dealer_by_amc[r.get("AMC")] = dealer_by_code[amcode]

    vm_header, vm_rows = source.find_table(["VMC", "VMN"])
    vendor_by_vmc = {r.get("VMC"): r for r in vm_rows} if vm_header else {}

    cm_header, cm_rows = source.find_table(["CMC", "CMN"])
    rto_by_cmc = {r.get("CMC"): _s(r.get("CMN")) for r in cm_rows} if cm_header else {}

    # The shared voucher table -- unique column combo so we match only it.
    inv_header, inv_rows = source.find_table(["VR", "CHASSIS", "IAMT", "TAXRATE", "SNAME"])
    inv_rows = inv_rows or []
    w_rows = [r for r in inv_rows if r.get("VR") == "W"]
    d_rows = [r for r in inv_rows if r.get("VR") == "D"]
    s_rows = [r for r in inv_rows if r.get("VR") == "S"]
    p_rows = [r for r in inv_rows if r.get("VR") == "P"]
    b_rows = [r for r in inv_rows if r.get("VR") == "B"]

    def _battery_fields(target, r):
        bm = battery_by_code.get(str(r.get("BMC") or ""))
        target.battery_maker = bm.name if bm else None
        target.battery_no1 = _batno(r.get("BAT1"))
        target.battery_no2 = _batno(r.get("BAT2"))
        target.battery_no3 = _batno(r.get("BAT3"))
        target.battery_no4 = _batno(r.get("BAT4"))

    def _accessory_fields(target, r):
        target.toolkit = _b(r.get("TOOLKIT"))
        target.jack = _b(r.get("JACK"))
        target.charger = _b(r.get("CHARGER"))
        target.mat = _b(r.get("MAT"))
        target.stapney = _b(r.get("STAPNEY"))
        target.front_glass = _b(r.get("FRONTGLASS"))
        target.h_lock = _b(r.get("HLOCK"))

    # --- Production Vouchers (VR='W') --------------------------------------
    # Same unique-constraint concern as the Delivery Challans/Tax Invoices
    # below, just on the other column: chassis_no is this row's own
    # identity (one row per physical vehicle), but the legacy data can have
    # two different chassis rows sharing one vou_no -- e.g. a voucher
    # number that was corrected/reissued. ProductionVoucher.vou_no is
    # unique, so writing it verbatim for every chassis can collide and
    # blow up the whole import on a single bad row.
    #
    # Resolved the same way as the sections below: a pure-Python pre-pass
    # decides which chassis "wins" each vou_no (preferring a non-cancelled
    # row, else first-encountered) before anything is added to the
    # session, so autoflush in the upsert pass has nothing ambiguous left
    # to race on. The losing chassis is still imported as its own vehicle
    # record -- it just keeps vou_no blank instead of failing the batch.
    parsed_w = []
    for r in w_rows:
        chassis = _s(r.get("CHASSIS"))
        if not chassis:
            continue
        parsed_w.append({
            "row": r,
            "chassis": chassis,
            "vou_no": _s(r.get("BNO")) or str(r.get("VNO") or ""),
            "cancelled": bool(r.get("DCHCANCEL")),
        })

    vou_winner = {}  # vou_no -> {"chassis": ..., "cancelled": ...}
    for p in parsed_w:
        vn = p["vou_no"]
        if not vn:
            continue
        cur = vou_winner.get(vn)
        if cur is None:
            vou_winner[vn] = {"chassis": p["chassis"], "cancelled": p["cancelled"]}
        elif p["cancelled"] and not cur["cancelled"]:
            pass  # this row loses -- current winner keeps the number
        elif cur["cancelled"] and not p["cancelled"]:
            vou_winner[vn] = {"chassis": p["chassis"], "cancelled": p["cancelled"]}
        # else: ambiguous duplicate -- first one already assigned wins.

    count = 0
    for p in parsed_w:
        r, chassis, vou_no = p["row"], p["chassis"], p["vou_no"]
        pv = ProductionVoucher.query.filter_by(chassis_no=chassis).first() or ProductionVoucher()
        wins_vou_no = bool(vou_no) and vou_winner.get(vou_no, {}).get("chassis") == chassis
        pv.vou_no = vou_no if wins_vou_no else None
        pv.date = _to_date(r.get("DT"))
        product = product_by_code.get(str(r.get("IMC") or ""))
        pv.product_name = product.name if product else "Unknown Product"
        pv.quantity = 1
        pv.chassis_no = chassis
        pv.motor_no = _s(r.get("MOTOR"))
        pv.controller_no = _s(r.get("CONTROLLER"))
        pv.differential_no = _s(r.get("DIFFERENTIAL"))
        colour_name = _s(r.get("COLOUR"))
        pv.colour = colour_name
        pv.colour_code = colour_code_by_name.get(colour_name) if colour_name else None
        pv.other = _s(r.get("OTHNO"))
        _battery_fields(pv, r)
        _accessory_fields(pv, r)
        remarks = " ".join(part for part in [_s(r.get("REM1")), _s(r.get("REM2"))] if part)
        pv.remarks = remarks or None
        db.session.add(pv)
        count += 1
    db.session.commit()
    summary["Production Vouchers"] = count

    # --- Delivery Challans (VR='D') -----------------------------------------
    # NOTE: DeliveryChallan.vehicle_id is unique -- only one challan can be
    # linked to a given vehicle. The legacy data sometimes has more than one
    # 'D' row for the same chassis (e.g. a cancelled challan re-issued under
    # a new number).
    #
    # Earlier versions of this import decided the winner while looping
    # through db.session.add()/query() calls for the same objects. That's
    # fragile: SQLAlchemy autoflush can fire mid-loop (any .first()/.filter
    # call flushes pending adds first), so a "loser" row's vehicle_id can
    # get written to the database before the winner has even been decided,
    # and two rows end up racing for one UNIQUE column.
    #
    # Fixed properly this time: resolve every vehicle's winning challan_no
    # in a pure-Python pass first (no db.session.add calls at all), then do
    # the actual upserts in a second pass where the outcome is already
    # fixed -- autoflush can fire freely there without any ambiguity left
    # to race on.

    # Pass 1: parse rows and resolve each vehicle's chassis (read-only
    # queries only -- nothing pending in the session yet).
    parsed_d = []
    for r in d_rows:
        chassis = _s(r.get("CHASSIS"))
        if not chassis:
            continue
        parsed_d.append({
            "row": r,
            "chassis": chassis,
            "challan_no": _s(r.get("BNO")) or str(r.get("VNO") or ""),
            "cancelled": bool(r.get("DCHCANCEL")),
            "vehicle": Vehicle.query.filter_by(chassis_no=chassis).first(),
        })

    # Pass 2: decide, purely in Python, which challan_no wins each vehicle
    # link. Prefer a non-cancelled row over a cancelled one; if it's still
    # ambiguous (both/neither cancelled), whichever is encountered first
    # in the source data keeps it.
    winner_by_vehicle = {}  # vehicle.id -> {"challan_no": ..., "cancelled": ...}
    for p in parsed_d:
        v = p["vehicle"]
        if not v:
            continue
        cur = winner_by_vehicle.get(v.id)
        if cur is None:
            winner_by_vehicle[v.id] = {"challan_no": p["challan_no"], "cancelled": p["cancelled"]}
        elif p["cancelled"] and not cur["cancelled"]:
            pass  # this row loses -- current winner keeps the link
        elif cur["cancelled"] and not p["cancelled"]:
            winner_by_vehicle[v.id] = {"challan_no": p["challan_no"], "cancelled": p["cancelled"]}
        # else: ambiguous duplicate -- first one already assigned wins.

    # Pass 3: upsert. The vehicle-link outcome is already fixed above, so
    # nothing here ever needs to re-open or reverse a decision -- safe to
    # let autoflush happen wherever SQLAlchemy wants.
    count = 0
    for p in parsed_d:
        r, chassis, challan_no, vehicle = p["row"], p["chassis"], p["challan_no"], p["vehicle"]
        dc = DeliveryChallan.query.filter_by(challan_no=challan_no).first() or DeliveryChallan()
        dc.challan_no = challan_no
        dc.date = _to_date(r.get("DT"))
        dc.cancelled = p["cancelled"]
        dealer = dealer_by_amc.get(r.get("AMC"))
        dc.dealer = dealer
        dc.destination = _s(r.get("REM1"))
        wins_vehicle = vehicle and winner_by_vehicle.get(vehicle.id, {}).get("challan_no") == challan_no
        dc.vehicle = vehicle if wins_vehicle else None
        product = product_by_code.get(str(r.get("IMC") or ""))
        dc.product_name = product.name if product else (vehicle.model_name if vehicle else None)
        dc.chassis_no = chassis
        dc.motor_no = _s(r.get("MOTOR")) or (vehicle.motor_no if vehicle else None)
        dc.controller_no = _s(r.get("CONTROLLER")) or (vehicle.controller_no if vehicle else None)
        dc.differential_no = _s(r.get("DIFFERENTIAL")) or (vehicle.differential_no if vehicle else None)
        dc.colour = _s(r.get("COLOUR")) or (vehicle.colour if vehicle else None)
        dc.other = _s(r.get("OTHNO")) or (vehicle.other if vehicle else None)
        _battery_fields(dc, r)
        _accessory_fields(dc, r)
        dc.sale_value = _to_float(r.get("NAMT"))
        # Salesman: prefer the value on this specific voucher row, falling
        # back to the dealer's default salesman (from the AMC master) --
        # this field was previously never set at all, so it always showed
        # blank on the Delivery Challan Register regardless of source data.
        dc.salesman = _s(r.get("SALESMAN")) or (dealer.salesman if dealer else None)
        dc.remarks1 = _s(r.get("REM1"))
        dc.remarks2 = _s(r.get("REM2"))
        db.session.add(dc)
        count += 1
        if vehicle:
            vehicle.stage = "Delivery Challan"
            if dealer:
                vehicle.dealer_name = dealer.name
    db.session.commit()
    summary["Delivery Challans"] = count

    # --- Tax Invoices (VR='S') ----------------------------------------------
    # Same unique-constraint concern as delivery challans above, on two
    # columns this time: TaxInvoice.vehicle_id AND
    # TaxInvoice.delivery_challan_id are both unique. Resolved the same way:
    # a pure-Python pre-pass decides every winner before any TaxInvoice is
    # added to the session, so autoflush in the upsert pass has nothing
    # ambiguous left to race on.

    # Pass 1: parse rows and look up their vehicle / matched delivery
    # challan (read-only queries; DeliveryChallan rows are already
    # committed from the block above, so this is safe regardless of order).
    parsed_s = []
    for r in s_rows:
        chassis = _s(r.get("CHASSIS"))
        parsed_s.append({
            "row": r,
            "chassis": chassis,
            "bill_no": _s(r.get("BNO")) or str(r.get("VNO") or ""),
            "cancelled": bool(r.get("DCHCANCEL")),
            "vehicle": Vehicle.query.filter_by(chassis_no=chassis).first() if chassis else None,
            "matched_dc": DeliveryChallan.query.filter_by(chassis_no=chassis).first() if chassis else None,
        })

    # Pass 2: resolve winners in Python. Same rule both times: prefer
    # non-cancelled over cancelled; otherwise first-encountered wins.
    def _resolve_winners(parsed, key_fn):
        winners = {}  # key -> {"bill_no": ..., "cancelled": ...}
        for p in parsed:
            k = key_fn(p)
            if k is None:
                continue
            cur = winners.get(k)
            if cur is None:
                winners[k] = {"bill_no": p["bill_no"], "cancelled": p["cancelled"]}
            elif p["cancelled"] and not cur["cancelled"]:
                pass
            elif cur["cancelled"] and not p["cancelled"]:
                winners[k] = {"bill_no": p["bill_no"], "cancelled": p["cancelled"]}
        return winners

    vehicle_winners = _resolve_winners(parsed_s, lambda p: p["vehicle"].id if p["vehicle"] else None)
    dc_winners = _resolve_winners(parsed_s, lambda p: p["matched_dc"].id if p["matched_dc"] else None)

    # Pass 3: upsert with the outcome already fixed.
    count = 0
    for p in parsed_s:
        r, chassis, bill_no = p["row"], p["chassis"], p["bill_no"]
        vehicle, matched_dc = p["vehicle"], p["matched_dc"]
        ti = TaxInvoice.query.filter_by(bill_no=bill_no).first() or TaxInvoice()
        ti.bill_no = bill_no
        ti.date = _to_date(r.get("DT"))
        ti.cancelled = p["cancelled"]

        wins_vehicle = vehicle and vehicle_winners.get(vehicle.id, {}).get("bill_no") == bill_no
        ti.vehicle = vehicle if wins_vehicle else None

        wins_dc = matched_dc and dc_winners.get(matched_dc.id, {}).get("bill_no") == bill_no
        ti.delivery_challan = matched_dc if wins_dc else None

        ti.buyer_name = _s(r.get("SNAME"))
        ti.buyer_relation, ti.buyer_father_name = _split_relation(r.get("SFATHER"))
        ti.buyer_address = ", ".join(p for p in [_s(r.get("SADD1")), _s(r.get("SADD2")), _s(r.get("SADD3"))] if p) or None
        ti.buyer_gst_no = _s(r.get("SGSTNO"))
        ti.buyer_aadhar = _s(r.get("SAADHAR"))
        ti.buyer_mobile = _s(r.get("SMOB"))
        ti.buyer_state = _s(r.get("SSTATE"))
        ti.buyer_state_code = _s(r.get("SSTCODE"))
        dealer = dealer_by_amc.get(r.get("AMC"))
        ti.dealer_name = dealer.name if dealer else None
        product = product_by_code.get(str(r.get("IMC") or ""))
        ti.product_name = product.name if product else (vehicle.model_name if vehicle else None)
        ti.chassis_no = chassis
        ti.motor_no = _s(r.get("MOTOR")) or (vehicle.motor_no if vehicle else None)
        ti.colour = _s(r.get("COLOUR")) or (vehicle.colour if vehicle else None)
        # Two distinct amounts on the legacy form: SALEAMT is the
        # "internal" sale amount, IAMT is the GST-taxable "Amount" the tax
        # is actually calculated on. Verified against the source data:
        # TAXAMT is consistently 5% (TAXRATE) of IAMT, and
        # IAMT + TAXAMT + INSAMT + REGAMT - DISAMT == NAMT exactly on every
        # row checked -- so NAMT is just the computed Bill Total, and IAMT
        # is the taxable base, not an "internal" figure. SALEAMT was
        # previously never read at all.
        ti.sale_amount = _to_float(r.get("SALEAMT"))
        ti.gst_sale_amount = _to_float(r.get("IAMT"))
        ti.discount = _to_float(r.get("DISAMT"))
        ti.gst_rate = _to_float(r.get("TAXRATE"))
        ti.insurance_amount = _to_float(r.get("INSAMT"))
        ti.registration_amount = _to_float(r.get("REGAMT"))
        financer = financer_by_code.get(str(r.get("HMC") or ""))
        ti.financer_name = financer.name if financer else None
        ti.hypothecation_amount = _to_float(r.get("HPAMT"))
        ti.amount_received = _to_float(r.get("SPAYRECD"))
        subsidy_raw = r.get("SUBSIDY")
        ti.subsidy_amount = _to_float(subsidy_raw) if isinstance(subsidy_raw, (int, float)) else 0.0
        ti.rto_name = rto_by_cmc.get(r.get("CMC"))
        ti.vehicle_reg_no = _s(r.get("VEHNO"))
        ti.despatch_through = _s(r.get("TRNO"))
        ti.eway_bill_no = _s(r.get("EWAYBILL"))
        ti.bank_name = _s(r.get("SCHQBANK"))
        ti.cancelled_cheque_no = _s(r.get("CHQNO")) or _s(r.get("SCHQ"))
        ti.remarks = _s(r.get("REM1"))
        ti.ledger_no = _s(r.get("LEDGER"))
        ti.voucher_no = _s(r.get("VOUNO"))
        ti.chassis_record_no = _s(r.get("CHASSISREC"))
        db.session.add(ti)
        count += 1
        if vehicle:
            vehicle.stage = "Tax Invoice"
    db.session.commit()
    summary["Tax Invoices"] = count

    # --- Purchase Bills (VR='P') --------------------------------------------
    # Several INV rows can share one bill (one row per raw-material line),
    # so group by the legacy voucher number (VNO) before writing headers.
    bills_by_vno = defaultdict(list)
    for r in p_rows:
        bills_by_vno[r.get("VNO")].append(r)
    count = 0
    for vno, items in bills_by_vno.items():
        first = items[0]
        bill_no = _s(first.get("BNO")) or str(vno)
        pb = PurchaseBill.query.filter_by(bill_no=bill_no).first() or PurchaseBill()
        vendor = vendor_by_vmc.get(first.get("VMC"))
        pb.bill_no = bill_no
        pb.date = _to_date(first.get("DT"))
        pb.party_name = _s(vendor.get("VMN")) if vendor else "Unknown Vendor"
        pb.party_gst_no = _s(vendor.get("GSTNO")) if vendor else None
        pb.party_state_code = _s(vendor.get("STCODE")) if vendor else "07"
        pb.remarks = _s(first.get("REM1"))
        db.session.add(pb)
        db.session.flush()  # need pb.id before writing items
        if pb.items:
            PurchaseBillItem.query.filter_by(bill_id=pb.id).delete()
        for it in items:
            product = product_by_code.get(str(it.get("IMC") or ""))
            db.session.add(PurchaseBillItem(
                bill_id=pb.id,
                item_name=product.name if product else "Unknown Item",
                hsn_code=product.hsn_code if product else None,
                qty=_to_float(it.get("WT")),
                rate=_to_float(it.get("RATE")),
                gst_rate=_to_float(it.get("TAXRATE")),
            ))
        count += 1
    db.session.commit()
    summary["Purchase Bills"] = count

    # --- Battery Delivery Challans (dedicated BDCH table + VR='B' rows) ----
    bdch_header, bdch_rows = source.find_table(["BNO", "BMC", "BAT1", "DCHCANCEL"])
    count = 0
    if bdch_header:
        for r in bdch_rows:
            challan_no = _s(r.get("BNO")) or str(r.get("VNO") or "")
            if not challan_no:
                continue
            bd = BatteryDeliveryChallan.query.filter_by(challan_no=challan_no).first() or BatteryDeliveryChallan()
            bd.challan_no = challan_no
            bd.date = _to_date(r.get("DT"))
            bd.dealer = dealer_by_amc.get(r.get("AMC"))
            bm = battery_by_code.get(str(r.get("BMC") or ""))
            bd.battery_maker = bm.name if bm else None
            bd.battery_no = _batno(r.get("BAT1")) or _batno(r.get("BAT2"))
            bd.qty = int(_to_float(r.get("WT")) or 1)
            bd.remarks = " ".join(p for p in [_s(r.get("REM1")), _s(r.get("REM2"))] if p) or None
            db.session.add(bd)
            count += 1
    # Legacy VR='B' rows come from the same combined voucher table as
    # everything else -- prefix their challan number so they can never
    # collide with the dedicated BDCH table's own numbering.
    for r in b_rows:
        challan_no = f"LEGACY-B-{r.get('VNO')}"
        bd = BatteryDeliveryChallan.query.filter_by(challan_no=challan_no).first() or BatteryDeliveryChallan()
        bd.challan_no = challan_no
        bd.date = _to_date(r.get("DT"))
        bd.dealer = dealer_by_amc.get(r.get("AMC"))
        bm = battery_by_code.get(str(r.get("BMC") or ""))
        bd.battery_maker = bm.name if bm else None
        bd.battery_no = _batno(r.get("BAT1")) or _batno(r.get("BAT2"))
        bd.qty = int(_to_float(r.get("WT")) or 1)
        chassis = _s(r.get("CHASSIS"))
        remarks = [p for p in [_s(r.get("REM1")), _s(r.get("REM2"))] if p]
        if chassis:
            remarks.append(f"Chassis: {chassis}")
        bd.remarks = " | ".join(remarks) or None
        db.session.add(bd)
        count += 1
    db.session.commit()
    summary["Battery Delivery Challans"] = count

    # --- Old Rickshaw / buy-back records ------------------------------------
    oh_header, oh_rows = source.find_table(["VEHNO", "ONAME", "SOLDAMT", "LOANAMT"])
    count = 0
    if oh_header:
        for r in oh_rows:
            vou_no = str(r.get("VNO") or "")
            reg_no = _s(r.get("VEHNO"))
            existing = (OldRickshaw.query.filter_by(vou_no=vou_no, vehicle_reg_no=reg_no).first()
                        if (vou_no or reg_no) else None)
            row = existing or OldRickshaw()
            row.vou_no = vou_no
            row.date = _to_date(r.get("DT"))
            dealer = dealer_by_amc.get(r.get("AMC"))
            row.party_name = dealer.name if dealer else None
            row.vehicle_reg_no = reg_no
            row.model_name = _s(r.get("IMN"))
            row.owner_name = _s(r.get("ONAME"))
            row.salesman = _s(r.get("SNAME"))
            row.sold_amount = _to_float(r.get("SOLDAMT"))
            row.loan_amount = _to_float(r.get("LOANAMT"))
            row.receipt_amount = _to_float(r.get("RCPTAMT"))
            row.receipt_no = _s(r.get("RNO"))
            row.ledger = _s(r.get("LEDGER"))
            row.resale_date = _to_date(r.get("SDT"))
            row.resale_ledger = _s(r.get("SLEDGER"))
            row.remarks1 = _s(r.get("REM1"))
            row.remarks2 = _s(r.get("REM2"))
            db.session.add(row)
            count += 1
        db.session.commit()
    summary["Old Rickshaw Records"] = count

    # --- Day Book Entries (legacy "FT" table, VR='C' rows) ------------------
    # FT holds two kinds of ledger rows, distinguished by VR:
    #   VR='C' -- plain cash/bank vouchers (manual Day Book entries: no
    #             chassis/product link, just party + amount + narration).
    #   VR='S' -- auto-posted entries generated FROM Tax Invoices (chassis
    #             and bill number populated) -- these mirror data already
    #             brought in by the Tax Invoice import above, so importing
    #             them again here would double-count amounts. Skipped.
    # DC marks direction: 'C' = credit (money received -> a receipt),
    # 'D' = debit (money paid out, or an opening balance owed).
    header, rows = source.find_table(["VR", "VNO", "AMC", "AMT", "DC", "NAR1"])
    count = 0
    if header:
        for r in rows:
            if r.get("VR") != "C":
                continue
            vno = r.get("VNO")
            vr_no = int(_to_float(vno)) if vno not in (None, "") else None
            dealer = dealer_by_amc.get(r.get("AMC"))
            existing = DayBook.query.filter_by(vr_no=vr_no).first() if vr_no else None
            row = existing or DayBook()
            row.vr_no = vr_no or DayBook.next_vr_no()
            row.date = _to_date(r.get("DT"))
            row.dealer_name = dealer.name if dealer else "Unknown Party"
            amt = _to_float(r.get("AMT"))
            dc = _s(r.get("DC"))
            row.credit_received = amt if dc == "C" else 0
            row.debit_paid = amt if dc == "D" else 0
            narration = " ".join(p for p in [
                _s(r.get("NAR1")), _s(r.get("NAR2")), _s(r.get("NAR3")), _s(r.get("NAR4")),
            ] if p)
            tno = r.get("TNO")
            if tno not in (None, 0, "0"):
                narration = f"{narration} (Txn #{tno})" if narration else f"Txn #{tno}"
            row.narration = narration or None
            db.session.add(row)
            count += 1
        db.session.commit()
    summary["Day Book Entries"] = count

    # --- Users (uid/pw/su/per) ----------------------------------------------
    header, rows = source.find_table(["uid", "pw", "su"])
    count = 0
    if header:
        for r in rows:
            username = r.get("uid")
            if not username:
                continue
            existing = User.query.filter_by(username=username).first()
            u = existing or User(username=username)
            if not existing:
                # NOTE: old plain-text password is used as the initial
                # hashed password so this legacy user can log in for the
                # first time; they should change it after that.
                #
                # Deliberately NOT done for an already-existing user: on a
                # re-import this would silently overwrite whatever password
                # is currently set -- including the app's own seeded
                # admin/admin1 bootstrap account if the legacy pw table
                # happens to also have a row named "admin", or any password
                # someone already changed by hand.
                u.set_password(str(r.get("pw") or "changeme"))
            u.is_super_user = (r.get("su") == "S")
            u.permissions = r.get("per")
            db.session.add(u)
            count += 1
        db.session.commit()
    summary["Users"] = count

    return summary


def run_import(filepath):
    """Import from an .xlsx export (same sheet layout as Book1.xlsx)."""
    return _run_from_source(ExcelSource(filepath))


def run_mdb_import(filepath, password=None):
    """
    Import directly from an Access .mdb/.accdb file. Tries two backends,
    in order:
      1. pyodbc + the Microsoft Access driver -- works on Windows, and is
         the only route that supports password-protected databases.
      2. mdbtools CLI -- works on Linux/Mac (or Windows under WSL). Has no
         way to supply a password, so it's skipped entirely when one is
         given.
    If neither is usable, raises MdbAccessUnavailable with a message that
    explains how to fix each path, plus the Excel-export fallback.
    """
    errors = []

    try:
        return _run_from_source(AccessOdbcSource(filepath, password=password))
    except AccessDriverMissing as e:
        errors.append(f"Windows/pyodbc route: {e}")

    try:
        return _run_from_source(MdbSource(filepath, password=password))
    except MdbToolsMissing as e:
        detail = str(e) if str(e) else "the mdb-tables/mdb-export commands aren't installed."
        errors.append(f"Linux/mdbtools route: {detail}")

    raise MdbAccessUnavailable(
        "Couldn't read this Access file with either available method.\n"
        + "\n".join(f"- {e}" for e in errors) + "\n\n"
        "Fixes:\n"
        "  Windows: pip install pyodbc, then install the free "
        "'Microsoft Access Database Engine Redistributable' from Microsoft "
        "(get the 32-bit build if your Python is 32-bit, 64-bit if it's 64-bit).\n"
        "  Linux/Mac: install mdbtools (apt-get install mdbtools, or brew install mdbtools).\n"
        "  Or simplest: open the file in Access and export the tables to Excel "
        "(same sheet/column names), then upload the .xlsx instead."
    )


if __name__ == "__main__":
    import sys
    from app import app
    with app.app_context():
        db.create_all()
        result = run_import(sys.argv[1] if len(sys.argv) > 1 else "data/Book1.xlsx")
        print(result)
