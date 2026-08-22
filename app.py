import os
from datetime import date, datetime as dt
from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, flash, session

from models import (db, Company, SimpleMaster, Dealer, Product, Vehicle, User,
                     ProductionFormula, ProductionVoucher, ProductionVoucherItem,
                     DeliveryChallan, TaxInvoice, PurchaseBill, PurchaseBillItem,
                     OldRickshaw, BatteryDeliveryChallan, JournalStock, DayBook)
from menu_config import MENU, find_item, all_items

BASE_DIR = os.path.abspath(os.path.dirname(__file__))

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "dev-secret-change-me")

# Database — reads DATABASE_URL from the environment so you can point this at
# PostgreSQL or MySQL in production, while still defaulting to a local SQLite
# file for quick testing with no setup required.
#
# Examples of DATABASE_URL:
#   PostgreSQL:  postgresql+psycopg2://ebill_user:mypassword@localhost:5432/ebill
#   MySQL:       mysql+pymysql://ebill_user:mypassword@localhost:3306/ebill
#   (unset)      sqlite:///.../ebill.db   <- local file, used automatically
# Production: set DATABASE_URL to the Supabase PostgreSQL connection string.
# Local development still falls back to SQLite if DATABASE_URL is not set.
db_url = os.environ.get("DATABASE_URL", f"sqlite:///{os.path.join(BASE_DIR, 'ebill.db')}")
# Some hosts hand out the old postgres:// prefix; SQLAlchemy expects postgresql://.
if db_url.startswith("postgres://"):
    db_url = db_url.replace("postgres://", "postgresql+psycopg2://", 1)
# Supabase connection strings normally already contain sslmode=require.
# If a Supabase URL is supplied without it, add SSL automatically.
if db_url.startswith("postgresql") and "supabase" in db_url and "sslmode=" not in db_url:
    db_url += ("&" if "?" in db_url else "?") + "sslmode=require"

app.config["SQLALCHEMY_DATABASE_URI"] = db_url
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
    "pool_pre_ping": True,
    "pool_recycle": 300,
}

db.init_app(app)


# ---------------------------------------------------------------------------
# Auth — mirrors the original desktop "Password" login dialog.
# Session-based; every route is protected except /login and static files.
# ---------------------------------------------------------------------------
@app.before_request
def _require_login():
    if request.endpoint in ("login", "static") or request.endpoint is None:
        return
    if "user_id" not in session:
        return redirect(url_for("login", next=request.path))


@app.route("/login", methods=["GET", "POST"])
def login():
    if "user_id" in session:
        return redirect(url_for("dashboard"))
    error = None
    userid = "admin"
    if request.method == "POST":
        userid = request.form.get("userid", "").strip()
        password = request.form.get("password", "")
        user = User.query.filter_by(username=userid).first()
        if user and user.check_password(password):
            session["user_id"] = user.id
            session["username"] = user.username
            session["is_super_user"] = user.is_super_user
            nxt = request.args.get("next") or url_for("dashboard")
            return redirect(nxt)
        error = "Invalid User ID or Password."
    return render_template("login.html", error=error, userid=userid)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


LOGOS_DIR = os.path.join(BASE_DIR, "static", "logos")
_LOGO_EXTENSIONS = [".jpg", ".jpeg", ".png", ".webp"]

HEADER_DIR = os.path.join(BASE_DIR, "static", "header")


def _find_model_logo(product):
    """
    Looks in static/logos/ for an image file named after the product's UMRN
    Code, then its Chassis Item Code, then the product name itself (first
    match wins) — case-insensitive, trying .jpg/.jpeg/.png/.webp in order.
    Returns a URL Flask can serve the image at.

    If no product-specific match is found, falls back to
    static/logos/_default.png (a placeholder) so the print preview always
    shows a logo slot rather than a blank gap -- drop the real logo files in
    static/logos/ (named per README.txt) to replace it per-product.
    """
    if not os.path.isdir(LOGOS_DIR):
        return None
    files_lower = {f.lower(): f for f in os.listdir(LOGOS_DIR)}
    candidates = [c for c in [product.umrn_code, product.chassis_item_code, product.name] if c] if product else []
    for code in candidates:
        for ext in _LOGO_EXTENSIONS:
            key = f"{code}{ext}".lower()
            if key in files_lower:
                return url_for("static", filename=f"logos/{files_lower[key]}")
    if "_default.png" in files_lower:
        return url_for("static", filename=f"logos/{files_lower['_default.png']}")
    return None


def _find_header_image(doc_key):
    """
    Looks in static/header/ for a full letterhead image to print at the top
    of a specific document type — e.g. doc_key="delivery_challan" looks for
    static/header/delivery_challan.jpg (or .jpeg/.png/.webp). Each document
    type has its own file, so this can be turned on for Delivery Challan
    without affecting Tax Invoice, Purchase Bill, etc. Returns a URL Flask
    can serve the image at, or None if no matching file exists (falls back
    to the built-in text header).
    """
    if not os.path.isdir(HEADER_DIR):
        return None
    files_lower = {f.lower(): f for f in os.listdir(HEADER_DIR)}
    for ext in _LOGO_EXTENSIONS:
        key = f"{doc_key}{ext}".lower()
        if key in files_lower:
            return url_for("static", filename=f"header/{files_lower[key]}")
    return None


def _parse_date(value):
    if not value:
        return dt.utcnow().date()
    try:
        return dt.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return dt.utcnow().date()


# AIS-007 (Revision 5), Table 11 — code for month and year of manufacture,
# embedded in the Chassis Number (VIN) at the 10th and 11th positions from
# the left. Skips "I", "O", and "Q" (not used, per the standard) — matches
# the letters actually listed in the table.
_AIS007_MONTH_CODE = {
    1: "A", 2: "B", 3: "C", 4: "D", 5: "E", 6: "F",
    7: "G", 8: "H", 9: "J", 10: "K", 11: "L", 12: "M",
}
_AIS007_YEAR_CODE = {
    2017: "A", 2018: "B", 2019: "C", 2020: "D", 2021: "E", 2022: "F",
    2023: "G", 2024: "H", 2025: "J", 2026: "K", 2027: "L", 2028: "M",
    2029: "N", 2030: "P", 2031: "R", 2032: "S", 2033: "T", 2034: "U",
    2035: "V", 2036: "W", 2037: "X", 2038: "Y", 2039: "Z", 2040: "1",
    2041: "2", 2042: "3", 2043: "4", 2044: "5", 2045: "6", 2046: "7",
}


def _ais007_month_year_code(the_date):
    """Returns (month_code, year_code) for a Production Voucher's date, per
    AIS-007 Table 11. Falls back to "?" for any year outside the table
    (i.e. before 2017 or after 2046) rather than guessing."""
    month_code = _AIS007_MONTH_CODE.get(the_date.month, "?")
    year_code = _AIS007_YEAR_CODE.get(the_date.year, "?")
    return month_code, year_code


def _export_excel(filename, headers, rows):
    """
    Generic 'Export Excel' helper shared by every report screen. Builds a
    simple .xlsx with a bold header row and auto-sized columns from a list
    of column headers and a list of row tuples/lists, and returns it as a
    downloadable response.
    """
    from io import BytesIO
    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(row)
    for i, header in enumerate(headers, start=1):
        max_len = max([len(str(header))] + [len(str(r[i - 1])) for r in rows if r[i - 1] is not None] or [0])
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(max_len + 4, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=filename,
                      mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.context_processor
def inject_globals():
    company = Company.query.first()
    return dict(
        MENU=MENU,
        company=company,
        today=date.today().strftime("%d/%m/%Y"),
        current_username=session.get("username"),
        current_is_super=session.get("is_super_user"),
    )


# ---------------------------------------------------------------------------
# Home / Dashboard  (mirrors the original launcher screen: left = full button
# grid of every module, right = "Dashboard as on <date>" pipeline table)
# ---------------------------------------------------------------------------
@app.route("/")
def dashboard():
    vehicles = Vehicle.query.order_by(Vehicle.date.desc()).limit(200).all()
    manufacturing = [v for v in vehicles if v.stage == "Manufacturing"]
    delivery_challan = [v for v in vehicles if v.stage == "Delivery Challan"]
    tax_invoice = [v for v in vehicles if v.stage == "Tax Invoice"]
    return render_template(
        "dashboard.html",
        manufacturing=manufacturing,
        delivery_challan=delivery_challan,
        tax_invoice=tax_invoice,
    )


# ---------------------------------------------------------------------------
# Generic module router — every menu item resolves to /m/<key> so navigation
# is 100% complete even for modules not fully built yet.
# ---------------------------------------------------------------------------
@app.route("/m/<key>", methods=["GET", "POST"])
def module(key):
    item = find_item(key)
    if item is None:
        return render_template("coming_soon.html", title=key, category=""), 404

    if item["kind"] == "custom":
        target = url_for(f"custom_{key.replace('-', '_')}")
        if request.query_string:
            target += "?" + request.query_string.decode()
        return redirect(target)

    if item["kind"] == "simple":
        return simple_master(item)

    # placeholder
    return render_template("coming_soon.html", title=item["label"], category=item["category"])


def simple_master(item):
    kind = item["key"]
    fields = item["fields"]

    if request.method == "POST":
        row_id = request.form.get("id")
        row = SimpleMaster.query.get(int(row_id)) if row_id else SimpleMaster(kind=kind)
        row.kind = kind
        row.name = request.form.get("name", "")
        row.code = request.form.get("code")
        row.address = request.form.get("address")
        row.mobile = request.form.get("mobile")
        row.account_no = request.form.get("account_no")
        row.ifsc = request.form.get("ifsc")
        row.extra = request.form.get("extra")
        has_default_field = any(f[0] == "is_default" for f in fields)
        if has_default_field:
            row.is_default = bool(request.form.get("is_default"))
            db.session.add(row)
            db.session.flush()  # get row.id before clearing others
            if row.is_default:
                # Only one record per kind can be the default (e.g. only one
                # primary Bank) — unset the flag on every other row.
                SimpleMaster.query.filter(
                    SimpleMaster.kind == kind, SimpleMaster.id != row.id
                ).update({"is_default": False})
        else:
            db.session.add(row)
        db.session.commit()
        flash(f"Saved '{row.name}'.", "success")
        return redirect(url_for("module", key=kind))

    rows = SimpleMaster.query.filter_by(kind=kind).order_by(SimpleMaster.name).all()
    return render_template("masters/simple_master.html", item=item, fields=fields, rows=rows)


@app.route("/m/<key>/delete/<int:row_id>", methods=["POST"])
def simple_master_delete(key, row_id):
    row = SimpleMaster.query.get_or_404(row_id)
    db.session.delete(row)
    db.session.commit()
    flash("Deleted.", "success")
    return redirect(url_for("module", key=key))


# ---------------------------------------------------------------------------
# Dealer Master — fully built reference CRUD module
# ---------------------------------------------------------------------------
@app.route("/m/dealer/full", methods=["GET", "POST"], endpoint="custom_dealer")
def dealer_master():
    if request.method == "POST":
        row_id = request.form.get("id")
        dealer = Dealer.query.get(int(row_id)) if row_id else Dealer()
        dealer.name = request.form.get("name", "")
        dealer.address1 = request.form.get("address1")
        dealer.address2 = request.form.get("address2")
        dealer.mobile = request.form.get("mobile")
        dealer.gst_no = request.form.get("gst_no")
        dealer.state = request.form.get("state")
        dealer.pan = request.form.get("pan")
        dealer.salesman = request.form.get("salesman")
        dealer.code = request.form.get("code") or None
        dealer.blocked = bool(request.form.get("blocked"))
        dealer.login_id = request.form.get("login_id")
        if request.form.get("password"):
            dealer.set_password(request.form.get("password"))
        db.session.add(dealer)
        db.session.commit()
        flash(f"Saved dealer '{dealer.name}'.", "success")
        return redirect(url_for("custom_dealer"))

    edit_id = request.args.get("edit", type=int)
    editing = Dealer.query.get(edit_id) if edit_id else None
    dealers = Dealer.query.order_by(Dealer.name).all()
    next_code_num = Dealer.query.count() + 1
    suggested_code = f"A-{next_code_num:02d}"
    return render_template("masters/dealer.html", dealers=dealers, editing=editing,
                           suggested_code=suggested_code)


@app.route("/m/dealer/delete/<int:dealer_id>", methods=["POST"])
def dealer_delete(dealer_id):
    d = Dealer.query.get_or_404(dealer_id)
    db.session.delete(d)
    db.session.commit()
    flash("Dealer deleted.", "success")
    return redirect(url_for("custom_dealer"))


# ---------------------------------------------------------------------------
# Product Master — fully built reference CRUD module
# ---------------------------------------------------------------------------
@app.route("/m/product/full", methods=["GET", "POST"], endpoint="custom_product")
def product_master():
    if request.method == "POST":
        row_id = request.form.get("id")
        p = Product.query.get(int(row_id)) if row_id else Product()
        p.name = (request.form.get("name") or "").strip()
        p.code = request.form.get("code") or None
        p.unit = request.form.get("unit") or "PCS"
        p.gst_rate = float(request.form.get("gst_rate") or 0)
        p.hsn_code = request.form.get("hsn_code")
        p.fro = request.form.get("fro") or "R"
        p.umrn_code = request.form.get("umrn_code") or None
        p.chassis_item_code = request.form.get("chassis_item_code") or None
        p.type_approval_no = request.form.get("type_approval_no") or None
        p.fuel_type = request.form.get("fuel_type") or "Battery/Electric"
        p.horn_db = request.form.get("horn_db") or None
        p.pass_by_db = request.form.get("pass_by_db") or None
        db.session.add(p)
        db.session.commit()
        flash(f"Saved product '{p.name}'.", "success")
        return redirect(url_for("custom_product"))

    edit_id = request.args.get("edit", type=int)
    editing = Product.query.get(edit_id) if edit_id else None
    products = Product.query.order_by(Product.name).all()
    products_json = [
        {
            "id": p.id, "name": p.name, "code": p.code, "unit": p.unit,
            "gst_rate": p.gst_rate, "hsn_code": p.hsn_code, "fro": p.fro,
            "umrn_code": p.umrn_code, "chassis_item_code": p.chassis_item_code,
            "type_approval_no": p.type_approval_no, "fuel_type": p.fuel_type,
            "horn_db": p.horn_db, "pass_by_db": p.pass_by_db,
        }
        for p in products
    ]
    return render_template("masters/product.html", products=products,
                           products_json=products_json, editing=editing)


@app.route("/m/product/delete/<int:product_id>", methods=["POST"])
def product_delete(product_id):
    p = Product.query.get_or_404(product_id)
    db.session.delete(p)
    db.session.commit()
    flash("Product deleted.", "success")
    return redirect(url_for("custom_product"))


# ---------------------------------------------------------------------------
# User Master — fully built reference CRUD module (passwords are hashed,
# unlike the original plain-text uid/pw table)
# ---------------------------------------------------------------------------
@app.route("/m/user/full", methods=["GET", "POST"], endpoint="custom_user")
def user_master():
    if request.method == "POST":
        row_id = request.form.get("id")
        u = User.query.get(int(row_id)) if row_id else User()
        u.username = request.form.get("username", "")
        if request.form.get("password"):
            u.set_password(request.form.get("password"))
        u.is_super_user = bool(request.form.get("is_super_user"))
        u.permissions = request.form.get("permissions")
        db.session.add(u)
        db.session.commit()
        flash(f"Saved user '{u.username}'.", "success")
        return redirect(url_for("custom_user"))

    users = User.query.order_by(User.username).all()
    return render_template("masters/user.html", users=users)


@app.route("/m/user/delete/<int:user_id>", methods=["POST"])
def user_delete(user_id):
    u = User.query.get_or_404(user_id)
    db.session.delete(u)
    db.session.commit()
    flash("User deleted.", "success")
    return redirect(url_for("custom_user"))


# ---------------------------------------------------------------------------
# Production Formula (BOM) management — Setup > 7
# Lets you build/edit the Bill of Material for each finished product by
# hand, instead of only getting it via Import Old Data. Grouped by product
# so you can add/edit/delete individual raw-material lines, or remove an
# entire formula at once.
# ---------------------------------------------------------------------------
@app.route("/m/production-formula/full", methods=["GET", "POST"], endpoint="custom_production_formula")
def production_formula():
    if request.method == "POST":
        row_id = request.form.get("id")
        formula_name = request.form.get("formula_name", "").strip()
        product_name = request.form.get("product_name", "").strip()
        raw_item_name = request.form.get("raw_item_name", "").strip()
        if not product_name or not raw_item_name:
            flash("Product and Raw Material are both required.", "error")
            return redirect(url_for("custom_production_formula"))
        # Formula Name defaults to the Finished Product name if left blank,
        # matching the legacy app's behaviour for pre-existing formulas.
        if not formula_name:
            formula_name = product_name

        row = ProductionFormula.query.get(int(row_id)) if row_id else ProductionFormula()
        row.formula_name = formula_name
        row.product_code = request.form.get("product_code") or None
        row.product_name = product_name
        row.raw_item_code = request.form.get("raw_item_code") or None
        row.raw_item_name = raw_item_name
        row.qty = float(request.form.get("qty") or 1)
        row.unit = request.form.get("unit") or "PCS"
        db.session.add(row)
        db.session.commit()
        flash(f"Saved formula line for '{formula_name}'.", "success")
        return redirect(url_for("custom_production_formula"))

    finished_products = Product.query.filter(
        (Product.fro == "F") | (Product.fro.is_(None))
    ).order_by(Product.name).all()
    raw_materials = Product.query.filter(
        (Product.fro == "R") | (Product.fro.is_(None))
    ).order_by(Product.name).all()

    lines = ProductionFormula.query.order_by(ProductionFormula.formula_name, ProductionFormula.raw_item_name).all()
    grouped = {}
    for line in lines:
        # Group by (Formula Name, Finished Product) — a product can have more
        # than one named formula, so product_name alone is not a unique key.
        key = (line.formula_name or line.product_name, line.product_name)
        grouped.setdefault(key, []).append(line)

    edit_id = request.args.get("edit", type=int)
    editing = ProductionFormula.query.get(edit_id) if edit_id else None

    return render_template(
        "setup/production_formula.html",
        grouped=grouped, finished_products=finished_products,
        raw_materials=raw_materials, editing=editing,
    )


@app.route("/m/production-formula/delete/<int:row_id>", methods=["POST"])
def production_formula_delete(row_id):
    row = ProductionFormula.query.get_or_404(row_id)
    db.session.delete(row)
    db.session.commit()
    flash("Formula line deleted.", "success")
    return redirect(url_for("custom_production_formula"))


@app.route("/m/production-formula/delete-product", methods=["POST"])
def production_formula_delete_product():
    formula_name = request.form.get("formula_name", "")
    product_name = request.form.get("product_name", "")
    ProductionFormula.query.filter_by(formula_name=formula_name, product_name=product_name).delete()
    db.session.commit()
    flash(f"Removed whole formula '{formula_name}'.", "success")
    return redirect(url_for("custom_production_formula"))


# ---------------------------------------------------------------------------
# User wise Option Setting — Setup > A
# Per-user checkbox grid of which menu items they're allowed to see/use.
# Super users are always unrestricted. Login/session now exists (see /login
# above); per-module enforcement using has_module_access() can be layered
# onto the `module` route below when needed.
# ---------------------------------------------------------------------------
@app.route("/m/option-setting/full", methods=["GET", "POST"], endpoint="custom_option_setting")
def option_setting():
    users = User.query.order_by(User.username).all()

    if request.method == "POST":
        user_id = request.form.get("user_id", type=int)
        u = User.query.get_or_404(user_id)
        selected = request.form.getlist("modules")
        u.allowed_modules = ",".join(selected)
        db.session.commit()
        flash(f"Module access updated for '{u.username}'.", "success")
        return redirect(url_for("custom_option_setting", user_id=u.id))

    selected_id = request.args.get("user_id", type=int) or (users[0].id if users else None)
    selected_user = User.query.get(selected_id) if selected_id else None
    selected_keys = set((selected_user.allowed_modules or "").split(",")) if selected_user and selected_user.allowed_modules else set()

    return render_template(
        "setup/option_setting.html",
        users=users, selected_user=selected_user, selected_keys=selected_keys,
    )


# ---------------------------------------------------------------------------
# Production Voucher — Vouchers > D
# Recording a finished vehicle coming off the line. Creates/updates the
# matching Vehicle (chassis) so it shows up on the Dashboard under
# "Manufacturing". If a Production Formula (BOM) exists for the chosen
# product, its raw-material lines are copied onto the voucher automatically.
# ---------------------------------------------------------------------------
@app.route("/m/production-voucher/full", methods=["GET"], endpoint="custom_production_voucher")
def production_voucher_list():
    vouchers = ProductionVoucher.query.order_by(
        ProductionVoucher.date.desc(), ProductionVoucher.id.desc()).all()
    finished_products = Product.query.filter(
        (Product.fro == "F") | (Product.fro.is_(None))
    ).order_by(Product.name).all()

    # "Formula Name / Finished Product" dropdown: one option per named
    # formula (Setup > 7), showing the Formula Name but carrying its actual
    # Finished Product underneath -- a product can have more than one named
    # formula, so each shows up as its own choice. Products with no formula
    # defined yet still get a plain option (by product name) so they can
    # still be manufactured; raw materials are just added by hand.
    formula_rows = ProductionFormula.query.order_by(ProductionFormula.formula_name).all()
    seen_formula_products = set()
    formula_options = []
    # Finished Product -> list of distinct Formula Names available for it
    # (a product can have more than one named formula/BOM).
    formulas_by_product = {}
    for f in formula_rows:
        key = (f.formula_name, f.product_name)
        if key in seen_formula_products:
            continue
        seen_formula_products.add(key)
        formula_options.append({"formula_name": f.formula_name, "product_name": f.product_name})
        formulas_by_product.setdefault(f.product_name, [])
        if f.formula_name not in formulas_by_product[f.product_name]:
            formulas_by_product[f.product_name].append(f.formula_name)
    products_with_formula = {f.product_name for f in formula_rows}
    unformulated_products = [p.name for p in finished_products if p.name not in products_with_formula]

    colours = SimpleMaster.query.filter_by(kind="colour").order_by(SimpleMaster.name).all()
    battery_makers = SimpleMaster.query.filter_by(kind="battery-maker").order_by(SimpleMaster.name).all()
    mechanics = SimpleMaster.query.filter_by(kind="mechanic").order_by(SimpleMaster.name).all()
    next_no = (db.session.query(db.func.max(ProductionVoucher.id)).scalar() or 0) + 1
    suggested_vou_no = f"W-{next_no + 17585}"
    return render_template(
        "vouchers/production_voucher.html",
        vouchers=vouchers, finished_products=finished_products,
        formula_options=formula_options, unformulated_products=unformulated_products,
        formulas_by_product=formulas_by_product,
        colours=colours, battery_makers=battery_makers, mechanics=mechanics,
        suggested_vou_no=suggested_vou_no, today_iso=date.today().isoformat(),
    )


@app.route("/m/production-voucher/formula-lines")
def production_voucher_formula_lines():
    """AJAX: raw-material lines for a Production Formula (BOM), used by the
    'Update From Formula' button. Prefers an exact Formula Name match (a
    product can have more than one named formula); falls back to matching
    by Finished Product name alone for products with no formula name set."""
    from flask import jsonify
    formula_name = request.args.get("formula", "").strip()
    product_name = request.args.get("product", "").strip()
    if formula_name:
        lines = ProductionFormula.query.filter_by(formula_name=formula_name, product_name=product_name).all()
    else:
        lines = ProductionFormula.query.filter_by(product_name=product_name).all()
    return jsonify([
        {"item_code": l.raw_item_code, "item_name": l.raw_item_name,
         "qty": l.qty, "unit": l.unit}
        for l in lines
    ])


@app.route("/m/production-voucher/generate-code")
def production_voucher_generate_code():
    """AJAX: suggest the next Chassis No. / Motor No. / Controller No. for a
    product, based on that Product's fixed 'Item Code' segment, the AIS-007
    (Rev 5) Table 11 month/year-of-manufacture code, and a running sequence.
    Mirrors the original 'Generate New Code' button.

    Chassis No. = <Item Code from Product/Item Master> + <month code> +
                  <year code> + <fixed 245/244 block> + <3-digit running
                  sequence, 001-999>.
    The 3-digit sequence wraps back to 001 after 999; each time it wraps,
    the fixed block alternates between 245 and 244 (so the two are never
    reused for the same sequence number back-to-back).

    Motor No. / Controller No. keep their own fixed prefixes but always end
    in the SAME 3-digit sequence number as the Chassis No. (this mirrors the
    legacy data, e.g. chassis '...245105' pairs with motor 'KTCM100000105' —
    both end in '105')."""
    from flask import jsonify
    product_name = request.args.get("product", "").strip()
    # If more than one Product row shares this name (e.g. one imported from
    # legacy data with no Item Code set, and a newer one edited via Modify
    # Product that DOES have it set), prefer whichever match actually has a
    # chassis_item_code configured rather than blindly taking the first row.
    product = (
        Product.query.filter(db.func.lower(Product.name) == product_name.lower(),
                              Product.chassis_item_code.isnot(None),
                              Product.chassis_item_code != "").first()
        or Product.query.filter(db.func.lower(Product.name) == product_name.lower()).first()
    )
    prefix = (product.chassis_item_code if product and product.chassis_item_code else "MD9GRDDE").strip()
    missing_item_code = not (product and product.chassis_item_code)

    voucher_date = _parse_date(request.args.get("date")) or date.today()
    month_code, year_code = _ais007_month_year_code(voucher_date)

    total = Vehicle.query.count()
    cycle = total // 999
    seq_num = (total % 999) + 1
    fixed_block = "245" if cycle % 2 == 0 else "244"
    seq_suffix = f"{seq_num:03d}"

    return jsonify({
        "chassis_no": f"{prefix}{month_code}{year_code}{fixed_block}{seq_suffix}",
        "motor_no": f"KTCM100000{seq_suffix}",
        "controller_no": f"KTCC100000{seq_suffix}",
        "month_code": month_code,
        "year_code": year_code,
        "missing_item_code": missing_item_code,
    })


@app.route("/m/production-voucher/new", methods=["POST"])
def production_voucher_create():
    product_name = request.form.get("product_name", "").strip()
    chassis_no = request.form.get("chassis_no", "").strip()
    vou_no = request.form.get("vou_no", "").strip()
    if not product_name or not chassis_no:
        flash("Product and Chassis No. are both required.", "error")
        return redirect(url_for("custom_production_voucher"))

    if Vehicle.query.filter_by(chassis_no=chassis_no).first():
        flash(f"Chassis No. '{chassis_no}' already exists.", "error")
        return redirect(url_for("custom_production_voucher"))

    if vou_no and ProductionVoucher.query.filter_by(vou_no=vou_no).first():
        flash(f"Vou. No. '{vou_no}' is already in use — please choose a different voucher number.", "error")
        return redirect(url_for("custom_production_voucher"))

    pv = ProductionVoucher(
        vou_no=vou_no,
        date=_parse_date(request.form.get("date")),
        product_name=product_name,
        quantity=int(request.form.get("quantity") or 1),
        chassis_no=chassis_no,
        motor_no=request.form.get("motor_no"),
        controller_no=request.form.get("controller_no"),
        differential_no=request.form.get("differential_no"),
        colour=request.form.get("colour"),
        colour_code=request.form.get("colour_code"),
        other=request.form.get("other"),
        battery_maker=request.form.get("battery_maker"),
        battery_no1=request.form.get("battery_no1"),
        battery_no2=request.form.get("battery_no2"),
        battery_no3=request.form.get("battery_no3"),
        battery_no4=request.form.get("battery_no4"),
        toolkit=bool(request.form.get("toolkit")),
        jack=bool(request.form.get("jack")),
        charger=bool(request.form.get("charger")),
        mat=bool(request.form.get("mat")),
        stapney=bool(request.form.get("stapney")),
        front_glass=bool(request.form.get("front_glass")),
        h_lock=bool(request.form.get("h_lock")),
        center_lock=bool(request.form.get("center_lock")),
        remarks=request.form.get("remarks"),
        machnic=request.form.get("machnic"),
    )
    db.session.add(pv)

    # Raw-material grid: the user can edit/add/remove lines in the popup
    # (pre-filled via "Update From Formula"). If they left the grid empty,
    # fall back to copying the BOM straight from Production Formula.
    item_names = request.form.getlist("item_name[]")
    item_codes = request.form.getlist("item_code[]")
    item_qtys = request.form.getlist("item_qty[]")
    item_units = request.form.getlist("item_unit[]")
    any_grid_rows = False
    for i, name in enumerate(item_names):
        name = (name or "").strip()
        if not name:
            continue
        any_grid_rows = True
        pv.items.append(ProductionVoucherItem(
            item_code=item_codes[i] if i < len(item_codes) else None,
            item_name=name,
            qty=float(item_qtys[i]) if i < len(item_qtys) and item_qtys[i] else 1,
            unit=item_units[i] if i < len(item_units) and item_units[i] else "PCS",
        ))
    if not any_grid_rows:
        formula_name = request.form.get("formula_name", "").strip()
        if formula_name:
            formula_lines = ProductionFormula.query.filter_by(
                formula_name=formula_name, product_name=product_name).all()
        else:
            formula_lines = ProductionFormula.query.filter_by(product_name=product_name).all()
        for fl in formula_lines:
            pv.items.append(ProductionVoucherItem(
                item_code=fl.raw_item_code, item_name=fl.raw_item_name,
                qty=fl.qty, unit=fl.unit,
            ))

    # Register the new chassis so it appears on the Dashboard
    vehicle = Vehicle(
        date=pv.date, model_name=product_name, chassis_no=chassis_no,
        motor_no=pv.motor_no, controller_no=pv.controller_no,
        differential_no=pv.differential_no, colour=pv.colour,
        colour_code=pv.colour_code, other=pv.other, stage="Manufacturing",
    )
    db.session.add(vehicle)

    db.session.commit()
    flash(f"Production Voucher {pv.vou_no} saved — chassis {chassis_no} added to Manufacturing.", "success")
    return redirect(url_for("custom_production_voucher"))


@app.route("/m/production-voucher/<int:voucher_id>/delete", methods=["POST"])
def production_voucher_delete(voucher_id):
    pv = ProductionVoucher.query.get_or_404(voucher_id)
    vehicle = Vehicle.query.filter_by(chassis_no=pv.chassis_no).first()
    if vehicle and vehicle.stage == "Manufacturing":
        db.session.delete(vehicle)
    db.session.delete(pv)
    db.session.commit()
    flash("Production Voucher deleted.", "success")
    return redirect(url_for("custom_production_voucher"))


# ---------------------------------------------------------------------------
# E-Rickshaw Delivery Challan — Vouchers > E
# Moves a manufactured chassis (Vehicle in "Manufacturing" stage) out to a
# dealer; updates the Dashboard pipeline automatically.
# ---------------------------------------------------------------------------
@app.route("/m/delivery-challan/full", methods=["GET"], endpoint="custom_delivery_challan")
def delivery_challan_list():
    challans = DeliveryChallan.query.order_by(
        DeliveryChallan.date.desc(), DeliveryChallan.id.desc()).all()

    # Bill No. per challan — prefer the actual Tax Invoice raised against it,
    # falling back to the challan's own Sale Bill No. snapshot if no invoice
    # exists yet. A challan is "Sold" if either one is filled in.
    invoiced_bill_no = {ti.delivery_challan_id: ti.bill_no
                         for ti in TaxInvoice.query.filter(TaxInvoice.delivery_challan_id.isnot(None)).all()}
    bill_no_by_challan = {}
    for c in challans:
        bill_no_by_challan[c.id] = invoiced_bill_no.get(c.id) or c.sale_bill_no or None

    available_vehicles = Vehicle.query.filter_by(stage="Manufacturing").order_by(Vehicle.chassis_no).all()
    dealers = Dealer.query.order_by(Dealer.name).all()
    battery_makers = SimpleMaster.query.filter_by(kind="battery-maker").order_by(SimpleMaster.name).all()
    next_no = (db.session.query(db.func.max(DeliveryChallan.id)).scalar() or 0) + 1
    suggested_challan_no = f"DC{next_no + 16500}"
    challans_by_id = {
        c.id: {
            "challan_no": c.challan_no or "", "date": c.date.isoformat() if c.date else "",
            "dealer_id": c.dealer_id or "", "dealer_name": c.dealer.name if c.dealer else "",
            "vehicle_id": c.vehicle_id or "",
            "product_name": c.product_name or "", "chassis_no": c.chassis_no or "",
            "motor_no": c.motor_no or "", "controller_no": c.controller_no or "", "colour": c.colour or "",
            "destination": c.destination or "",
            "battery_maker": c.battery_maker or "", "battery_no1": c.battery_no1 or "",
            "battery_no2": c.battery_no2 or "", "battery_no3": c.battery_no3 or "", "battery_no4": c.battery_no4 or "",
            "toolkit": c.toolkit, "jack": c.jack, "charger": c.charger, "mat": c.mat,
            "stapney": c.stapney, "front_glass": c.front_glass, "center_lock": c.center_lock, "h_lock": c.h_lock,
            "salesman": c.salesman or "", "sale_bill_no": c.sale_bill_no or "", "sale_value": c.sale_value or 0,
            "remarks1": c.remarks1 or "", "remarks2": c.remarks2 or "",
            "invoiced": c.id in invoiced_bill_no,
        }
        for c in challans
    }
    return render_template(
        "vouchers/delivery_challan.html",
        challans=challans, available_vehicles=available_vehicles,
        dealers=dealers, battery_makers=battery_makers,
        suggested_challan_no=suggested_challan_no, today_iso=date.today().isoformat(),
        bill_no_by_challan=bill_no_by_challan, challans_by_id=challans_by_id,
    )


@app.route("/m/delivery-challan/new", methods=["POST"])
def delivery_challan_create():
    vehicle_id = request.form.get("vehicle_id", type=int)
    dealer_id = request.form.get("dealer_id", type=int)
    if not vehicle_id or not dealer_id:
        flash("Chassis and Dealer are both required.", "error")
        return redirect(url_for("custom_delivery_challan"))

    vehicle = Vehicle.query.get_or_404(vehicle_id)
    dealer = Dealer.query.get_or_404(dealer_id)

    if vehicle.stage != "Manufacturing":
        flash("That chassis has already moved past Manufacturing.", "error")
        return redirect(url_for("custom_delivery_challan"))

    dc = DeliveryChallan(
        challan_no=request.form.get("challan_no"),
        date=_parse_date(request.form.get("date")),
        dealer_id=dealer.id,
        vehicle_id=vehicle.id,
        destination=request.form.get("destination"),
        product_name=vehicle.model_name,
        chassis_no=vehicle.chassis_no,
        motor_no=vehicle.motor_no,
        controller_no=vehicle.controller_no,
        differential_no=vehicle.differential_no,
        colour=vehicle.colour,
        other=vehicle.other,
        battery_maker=request.form.get("battery_maker"),
        battery_no1=request.form.get("battery_no1"),
        battery_no2=request.form.get("battery_no2"),
        battery_no3=request.form.get("battery_no3"),
        battery_no4=request.form.get("battery_no4"),
        toolkit=bool(request.form.get("toolkit")),
        jack=bool(request.form.get("jack")),
        charger=bool(request.form.get("charger")),
        center_lock=bool(request.form.get("center_lock")),
        mat=bool(request.form.get("mat")),
        stapney=bool(request.form.get("stapney")),
        front_glass=bool(request.form.get("front_glass")),
        h_lock=bool(request.form.get("h_lock")),
        salesman=request.form.get("salesman"),
        sale_bill_no=request.form.get("sale_bill_no"),
        sale_value=float(request.form.get("sale_value") or 0),
        remarks1=request.form.get("remarks1"),
        remarks2=request.form.get("remarks2"),
    )
    db.session.add(dc)

    # Move the chassis forward in the pipeline
    vehicle.stage = "Delivery Challan"
    vehicle.dealer_name = dealer.name
    db.session.add(vehicle)

    db.session.commit()
    flash(f"Delivery Challan {dc.challan_no} saved for chassis {vehicle.chassis_no}.", "success")
    return redirect(url_for("custom_delivery_challan"))


@app.route("/m/delivery-challan/<int:challan_id>/update", methods=["POST"], endpoint="delivery_challan_update")
def delivery_challan_update(challan_id):
    """Modify an existing Delivery Challan. The linked chassis/dealer stay
    fixed (changing them would desync the Dashboard pipeline) — only the
    challan's own details are editable, same idea as Tax Invoice's Modify."""
    dc = DeliveryChallan.query.get_or_404(challan_id)

    # Dealer / Chassis are editable, but only while no Tax Invoice has been
    # raised against this challan yet — once invoiced, the chassis/dealer
    # are baked into that invoice's own snapshot, so changing them here
    # would desync the two records. Cancel and re-create instead in that case.
    new_dealer_id = request.form.get("dealer_id", type=int)
    new_vehicle_id = request.form.get("vehicle_id", type=int)
    changing_link = bool(new_dealer_id) and bool(new_vehicle_id) and (
        new_dealer_id != dc.dealer_id or new_vehicle_id != dc.vehicle_id)

    if changing_link:
        has_invoice = TaxInvoice.query.filter_by(delivery_challan_id=dc.id).first() is not None
        if has_invoice:
            flash("Dealer/Chassis can't be changed — a Tax Invoice already exists for this Delivery Challan. "
                  "Cancel and re-create it instead if that's wrong.", "error")
        else:
            new_vehicle = Vehicle.query.get_or_404(new_vehicle_id)
            new_dealer = Dealer.query.get_or_404(new_dealer_id)
            if new_vehicle.id != dc.vehicle_id and new_vehicle.stage != "Manufacturing":
                flash("That chassis has already moved past Manufacturing.", "error")
            else:
                # Free up the old chassis if it's being swapped out
                if dc.vehicle and dc.vehicle.id != new_vehicle.id:
                    dc.vehicle.stage = "Manufacturing"
                    dc.vehicle.dealer_name = None
                dc.dealer_id = new_dealer.id
                dc.vehicle_id = new_vehicle.id
                dc.product_name = new_vehicle.model_name
                dc.chassis_no = new_vehicle.chassis_no
                dc.motor_no = new_vehicle.motor_no
                dc.controller_no = new_vehicle.controller_no
                dc.differential_no = new_vehicle.differential_no
                dc.colour = new_vehicle.colour
                dc.other = new_vehicle.other
                new_vehicle.stage = "Delivery Challan"
                new_vehicle.dealer_name = new_dealer.name

    dc.challan_no = request.form.get("challan_no") or dc.challan_no
    dc.date = _parse_date(request.form.get("date")) or dc.date
    dc.destination = request.form.get("destination")
    dc.battery_maker = request.form.get("battery_maker")
    dc.battery_no1 = request.form.get("battery_no1")
    dc.battery_no2 = request.form.get("battery_no2")
    dc.battery_no3 = request.form.get("battery_no3")
    dc.battery_no4 = request.form.get("battery_no4")
    dc.toolkit = bool(request.form.get("toolkit"))
    dc.jack = bool(request.form.get("jack"))
    dc.charger = bool(request.form.get("charger"))
    dc.center_lock = bool(request.form.get("center_lock"))
    dc.mat = bool(request.form.get("mat"))
    dc.stapney = bool(request.form.get("stapney"))
    dc.front_glass = bool(request.form.get("front_glass"))
    dc.h_lock = bool(request.form.get("h_lock"))
    dc.salesman = request.form.get("salesman")
    dc.sale_bill_no = request.form.get("sale_bill_no")
    dc.sale_value = float(request.form.get("sale_value") or 0)
    dc.remarks1 = request.form.get("remarks1")
    dc.remarks2 = request.form.get("remarks2")

    db.session.commit()
    flash(f"Delivery Challan {dc.challan_no} updated.", "success")
    return redirect(url_for("custom_delivery_challan"))


@app.route("/m/delivery-challan/<int:challan_id>/cancel", methods=["POST"])
def delivery_challan_cancel(challan_id):
    dc = DeliveryChallan.query.get_or_404(challan_id)
    dc.cancelled = not dc.cancelled
    # cancelling sends the chassis back to Manufacturing so it's available again
    if dc.vehicle:
        dc.vehicle.stage = "Manufacturing" if dc.cancelled else "Delivery Challan"
        dc.vehicle.dealer_name = None if dc.cancelled else dc.vehicle.dealer_name
    db.session.commit()
    flash("Challan " + ("cancelled" if dc.cancelled else "un-cancelled") + ".", "success")
    return redirect(url_for("custom_delivery_challan"))


@app.route("/m/delivery-challan/<int:challan_id>/delete", methods=["POST"])
def delivery_challan_delete(challan_id):
    dc = DeliveryChallan.query.get_or_404(challan_id)
    if dc.vehicle:
        dc.vehicle.stage = "Manufacturing"
        dc.vehicle.dealer_name = None
    db.session.delete(dc)
    db.session.commit()
    flash("Delivery Challan deleted; chassis returned to Manufacturing.", "success")
    return redirect(url_for("custom_delivery_challan"))


@app.route("/m/delivery-challan/<int:challan_id>/print")
def delivery_challan_print(challan_id):
    dc = DeliveryChallan.query.get_or_404(challan_id)
    company = Company.query.first()
    product = Product.query.filter_by(name=dc.product_name).first()
    model_logo_url = _find_model_logo(product)
    header_image_url = _find_header_image("delivery_challan")
    return render_template("vouchers/delivery_challan_print.html", dc=dc, company=company,
                            model_logo_url=model_logo_url, header_image_url=header_image_url)


# ---------------------------------------------------------------------------
# Tax Invoice — Vouchers > F
# The GST sale invoice raised against an already-delivered chassis. Splits
# CGST+SGST vs IGST automatically based on the buyer's state code, same as
# the original software.
# ---------------------------------------------------------------------------
@app.route("/m/tax-invoice/full", methods=["GET"], endpoint="custom_tax_invoice")
def tax_invoice_list():
    invoices = TaxInvoice.query.order_by(TaxInvoice.date.desc(), TaxInvoice.id.desc()).all()
    # only challans that don't already have an invoice
    invoiced_challan_ids = {ti.delivery_challan_id for ti in TaxInvoice.query.all() if ti.delivery_challan_id}
    available_challans = [c for c in DeliveryChallan.query.filter_by(cancelled=False).order_by(DeliveryChallan.date.desc()).all()
                           if c.id not in invoiced_challan_ids]
    financers = SimpleMaster.query.filter_by(kind="financer").order_by(SimpleMaster.name).all()
    rtos = SimpleMaster.query.filter_by(kind="rto").order_by(SimpleMaster.name).all()
    # "F3 : Customer List" on the Invoice form (legacy desktop app) — no
    # separate Customer master exists yet, so this reuses Dealer Master
    # records (buyer already defaults to the dealer) to auto-fill address/
    # GST/mobile/state when a saved name is picked.
    customer_list = Dealer.query.order_by(Dealer.name).all()
    # Bank field is driven by the Bank Details master (Setup > B) so the
    # Account No. / IFSC printed on the invoice always match a saved bank.
    banks = SimpleMaster.query.filter_by(kind="bank").order_by(SimpleMaster.name).all()
    next_no = (db.session.query(db.func.max(TaxInvoice.id)).scalar() or 0) + 1
    suggested_bill_no = f"GRD/{next_no + 1000}/26-27"

    # Editable-field snapshot for each invoice, used by the "Modify" /
    # double-click-to-edit JS on the Invoice List to fill the popup without
    # an extra round-trip.
    edit_fields = [
        "bill_no", "buyer_name", "buyer_relation", "buyer_father_name",
        "buyer_address", "buyer_gst_no", "buyer_pan",
        "buyer_aadhar", "buyer_mobile", "buyer_state", "buyer_state_code", "state_type",
        "dealer_name", "product_name", "colour", "chassis_no", "motor_no", "controller_no", "other_desc",
        "sale_amount", "gst_sale_amount", "discount", "gst_rate",
        "insurance_amount", "registration_amount", "financer_name",
        "hypothecation_amount", "amount_received", "subsidy_amount",
        "rto_name", "vehicle_reg_no", "despatch_through", "eway_bill_no",
        "mode_term", "bank_name", "bank_ifsc", "cvr_no", "license_no", "cancelled_cheque_no", "remarks", "ledger_no",
        "voucher_no", "chassis_record_no",
    ]
    invoices_by_id = {
        ti.id: {
            **{f: (getattr(ti, f) if getattr(ti, f) is not None else "") for f in edit_fields},
            "date": ti.date.isoformat() if ti.date else "",
            "buyer_dob": ti.buyer_dob.isoformat() if ti.buyer_dob else "",
        }
        for ti in invoices
    }

    return render_template(
        "vouchers/tax_invoice.html",
        invoices=invoices, available_challans=available_challans,
        financers=financers, rtos=rtos, banks=banks, customer_list=customer_list,
        suggested_bill_no=suggested_bill_no, today_iso=date.today().isoformat(),
        invoices_by_id=invoices_by_id,
    )


@app.route("/m/tax-invoice/new", methods=["POST"])
def tax_invoice_create():
    challan_id = request.form.get("challan_id", type=int)
    if not challan_id:
        flash("Please choose a Delivery Challan to invoice.", "error")
        return redirect(url_for("custom_tax_invoice"))

    challan = DeliveryChallan.query.get_or_404(challan_id)
    if TaxInvoice.query.filter_by(delivery_challan_id=challan.id).first():
        flash("That Delivery Challan already has a Tax Invoice.", "error")
        return redirect(url_for("custom_tax_invoice"))

    product = Product.query.filter_by(name=challan.product_name).first()
    default_gst = product.gst_rate if product else 5
    buyer_name = request.form.get("buyer_name") or (challan.dealer.name if challan.dealer else "")

    # Bank is the customer's bank (for hypothecation/loan details) and is
    # freely editable — typing a name that matches the Bank Details master
    # (Setup > B) still auto-fills Account No. / IFSC from there, but any
    # other name (or a hand-typed IFSC) is kept as entered. If the field
    # was left blank entirely, fall back to whichever Bank is marked
    # Default / Primary.
    bank_name = request.form.get("bank_name", "").strip()
    bank_ifsc_input = request.form.get("bank_ifsc", "").strip()
    bank_master = SimpleMaster.query.filter_by(kind="bank", name=bank_name).first() if bank_name else None
    if not bank_name:
        bank_master = SimpleMaster.query.filter_by(kind="bank", is_default=True).first()
        if bank_master:
            bank_name = bank_master.name
    bank_account_no = bank_master.account_no if bank_master else None
    bank_ifsc = bank_ifsc_input or (bank_master.ifsc if bank_master else None)

    ti = TaxInvoice(
        bill_no=request.form.get("bill_no"),
        date=_parse_date(request.form.get("date")),
        delivery_challan_id=challan.id,
        vehicle_id=challan.vehicle_id,
        buyer_name=buyer_name,
        buyer_relation=request.form.get("buyer_relation") or "S/o",
        buyer_father_name=request.form.get("buyer_father_name"),
        buyer_address=request.form.get("buyer_address"),
        buyer_gst_no=request.form.get("buyer_gst_no"),
        buyer_pan=request.form.get("buyer_pan"),
        buyer_aadhar=request.form.get("buyer_aadhar"),
        buyer_mobile=request.form.get("buyer_mobile"),
        buyer_state=request.form.get("buyer_state"),
        buyer_state_code=request.form.get("buyer_state_code"),
        state_type=request.form.get("state_type") or "I",
        buyer_dob=_parse_date(request.form.get("buyer_dob")),
        dealer_name=challan.dealer.name if challan.dealer else None,
        product_name=challan.product_name,
        chassis_no=challan.chassis_no,
        motor_no=challan.motor_no,
        controller_no=challan.controller_no,
        other_desc=challan.other,
        colour=challan.colour,
        sale_amount=float(request.form.get("sale_amount") or challan.sale_value or 0),
        gst_sale_amount=float(request.form.get("gst_sale_amount")
                               or request.form.get("sale_amount") or challan.sale_value or 0),
        discount=float(request.form.get("discount") or 0),
        gst_rate=float(request.form.get("gst_rate") or default_gst or 5),
        insurance_amount=float(request.form.get("insurance_amount") or 0),
        registration_amount=float(request.form.get("registration_amount") or 0),
        financer_name=request.form.get("financer_name"),
        hypothecation_amount=float(request.form.get("hypothecation_amount") or 0),
        amount_received=float(request.form.get("amount_received") or 0),
        subsidy_amount=float(request.form.get("subsidy_amount") or 0),
        rto_name=request.form.get("rto_name"),
        vehicle_reg_no=request.form.get("vehicle_reg_no"),
        despatch_through=request.form.get("despatch_through"),
        eway_bill_no=request.form.get("eway_bill_no"),
        mode_term=request.form.get("mode_term") or "BANK/CASH",
        bank_name=bank_name,
        bank_account_no=bank_account_no,
        bank_ifsc=bank_ifsc,
        cvr_no=request.form.get("cvr_no"),
        license_no=request.form.get("license_no"),
        cancelled_cheque_no=request.form.get("cancelled_cheque_no"),
        remarks=request.form.get("remarks"),
        ledger_no=request.form.get("ledger_no"),
        voucher_no=request.form.get("voucher_no"),
        chassis_record_no=request.form.get("chassis_record_no"),
    )
    db.session.add(ti)

    if challan.vehicle:
        challan.vehicle.stage = "Tax Invoice"

    db.session.commit()
    flash(f"Tax Invoice {ti.bill_no} saved for chassis {ti.chassis_no}.", "success")
    return redirect(url_for("tax_invoice_print_options", invoice_id=ti.id))


@app.route("/m/tax-invoice/<int:invoice_id>/update", methods=["POST"], endpoint="tax_invoice_update")
def tax_invoice_update(invoice_id):
    """Modify (double-click / toolbar Modify on the Invoice List). Edits the
    buyer/amount/RTO/financer/bank details of an already-saved invoice — the
    linked chassis (Delivery Challan / Vehicle) itself is not changed here."""
    ti = TaxInvoice.query.get_or_404(invoice_id)

    bank_name = request.form.get("bank_name", "").strip()
    bank_ifsc_input = request.form.get("bank_ifsc", "").strip()
    bank_master = SimpleMaster.query.filter_by(kind="bank", name=bank_name).first() if bank_name else None
    if not bank_name:
        bank_master = SimpleMaster.query.filter_by(kind="bank", is_default=True).first()
        if bank_master:
            bank_name = bank_master.name
    bank_account_no = bank_master.account_no if bank_master else None
    bank_ifsc = bank_ifsc_input or (bank_master.ifsc if bank_master else None)

    ti.bill_no = request.form.get("bill_no") or ti.bill_no
    ti.date = _parse_date(request.form.get("date")) or ti.date
    ti.buyer_name = request.form.get("buyer_name")
    ti.buyer_relation = request.form.get("buyer_relation") or "S/o"
    ti.buyer_father_name = request.form.get("buyer_father_name")
    ti.buyer_address = request.form.get("buyer_address")
    ti.buyer_gst_no = request.form.get("buyer_gst_no")
    ti.buyer_pan = request.form.get("buyer_pan")
    ti.buyer_aadhar = request.form.get("buyer_aadhar")
    ti.buyer_mobile = request.form.get("buyer_mobile")
    ti.buyer_state = request.form.get("buyer_state")
    ti.buyer_state_code = request.form.get("buyer_state_code")
    ti.state_type = request.form.get("state_type") or "I"
    ti.buyer_dob = _parse_date(request.form.get("buyer_dob")) or ti.buyer_dob
    ti.controller_no = request.form.get("controller_no") or ti.controller_no
    ti.other_desc = request.form.get("other_desc") or ti.other_desc
    ti.sale_amount = float(request.form.get("sale_amount") or 0)
    ti.gst_sale_amount = float(request.form.get("gst_sale_amount") or request.form.get("sale_amount") or 0)
    ti.discount = float(request.form.get("discount") or 0)
    ti.gst_rate = float(request.form.get("gst_rate") or ti.gst_rate or 5)
    ti.insurance_amount = float(request.form.get("insurance_amount") or 0)
    ti.registration_amount = float(request.form.get("registration_amount") or 0)
    ti.financer_name = request.form.get("financer_name")
    ti.hypothecation_amount = float(request.form.get("hypothecation_amount") or 0)
    ti.amount_received = float(request.form.get("amount_received") or 0)
    ti.subsidy_amount = float(request.form.get("subsidy_amount") or 0)
    ti.rto_name = request.form.get("rto_name")
    ti.vehicle_reg_no = request.form.get("vehicle_reg_no")
    ti.despatch_through = request.form.get("despatch_through")
    ti.eway_bill_no = request.form.get("eway_bill_no")
    ti.mode_term = request.form.get("mode_term") or "BANK/CASH"
    ti.bank_name = bank_name
    ti.bank_account_no = bank_account_no
    ti.bank_ifsc = bank_ifsc
    ti.cvr_no = request.form.get("cvr_no")
    ti.license_no = request.form.get("license_no")
    ti.cancelled_cheque_no = request.form.get("cancelled_cheque_no")
    ti.remarks = request.form.get("remarks")
    ti.ledger_no = request.form.get("ledger_no")
    ti.voucher_no = request.form.get("voucher_no")
    ti.chassis_record_no = request.form.get("chassis_record_no")

    db.session.commit()
    flash(f"Tax Invoice {ti.bill_no} updated.", "success")
    return redirect(url_for("custom_tax_invoice"))


@app.route("/m/tax-invoice/<int:invoice_id>/cancel", methods=["POST"])
def tax_invoice_cancel(invoice_id):
    ti = TaxInvoice.query.get_or_404(invoice_id)
    ti.cancelled = not ti.cancelled
    if ti.vehicle:
        ti.vehicle.stage = "Delivery Challan" if ti.cancelled else "Tax Invoice"
    db.session.commit()
    flash("Invoice " + ("cancelled" if ti.cancelled else "un-cancelled") + ".", "success")
    return redirect(url_for("custom_tax_invoice"))


@app.route("/m/tax-invoice/<int:invoice_id>/delete", methods=["POST"])
def tax_invoice_delete(invoice_id):
    ti = TaxInvoice.query.get_or_404(invoice_id)
    if ti.vehicle:
        ti.vehicle.stage = "Delivery Challan"
    db.session.delete(ti)
    db.session.commit()
    flash("Tax Invoice deleted; chassis returned to Delivery Challan.", "success")
    return redirect(url_for("custom_tax_invoice"))


@app.route("/m/tax-invoice/<int:invoice_id>/update-payment", methods=["POST"])
def tax_invoice_update_payment(invoice_id):
    """Payment Details popup — Payment Rec'able Report. Lets you fill in
    everything the desktop popup shows without re-opening the full Tax
    Invoice form: amounts, RTO/legal reference numbers, and the subsidy
    claim status."""
    ti = TaxInvoice.query.get_or_404(invoice_id)
    ti.sale_amount = float(request.form.get("sale_amount") or ti.sale_amount or 0)
    ti.hypothecation_amount = float(request.form.get("hypothecation_amount") or 0)
    ti.amount_received = float(request.form.get("amount_received") or 0)
    ti.subsidy_amount = float(request.form.get("subsidy_amount") or ti.subsidy_amount or 0)
    ti.subsidy_status = request.form.get("subsidy_status") or ti.subsidy_status or "Due"
    ti.financer_name = request.form.get("financer_name") or ti.financer_name
    ti.voucher_no = request.form.get("voucher_no")
    ti.chassis_record_no = request.form.get("chassis_record_no")
    ti.ledger_no = request.form.get("ledger_no")
    ti.cancelled_cheque_no = request.form.get("cancelled_cheque_no")
    ti.vehicle_reg_no = request.form.get("vehicle_reg_no")
    db.session.commit()
    flash(f"Payment updated for {ti.bill_no}.", "success")
    return redirect(request.form.get("return_to") or url_for("custom_payment_receivable_report"))


DOC_TYPES = {
    "invoice":        {"title": "Tax Invoice",          "no_label": "Invoice No.",   "endpoint": "tax_invoice_print",        "ready": True},
    "proforma":       {"title": "Proforma Invoice",      "no_label": "Proforma No.",  "endpoint": "tax_invoice_print",        "ready": True},
    "manufacturing":  {"title": "Manufacturing Invoice", "no_label": "Invoice No.",   "endpoint": "tax_invoice_print",        "ready": True},
    "affidavit":      {"title": "Affidavit",             "no_label": "Ref No.",       "endpoint": "tax_invoice_affidavit",    "ready": True},
    "undertaking":    {"title": "Undertaking",           "no_label": "Ref No.",       "endpoint": "tax_invoice_undertaking",  "ready": True},
    "form22":         {"title": "Form 22",               "no_label": "Ref No.",       "endpoint": "tax_invoice_form22",       "ready": True},
}


@app.route("/m/tax-invoice/<int:invoice_id>/print-options", endpoint="tax_invoice_print_options")
def tax_invoice_print_options(invoice_id):
    ti = TaxInvoice.query.get_or_404(invoice_id)
    return render_template("vouchers/tax_invoice_print_options.html", ti=ti, doc_types=DOC_TYPES)


@app.route("/m/tax-invoice/<int:invoice_id>/print")
def tax_invoice_print(invoice_id):
    ti = TaxInvoice.query.get_or_404(invoice_id)
    company = Company.query.first()
    doc = request.args.get("doc", "invoice")
    meta = DOC_TYPES.get(doc, DOC_TYPES["invoice"])
    if not meta["ready"]:
        flash(f"{meta['title']} format hasn't been set up yet — send a sample printout/format and it'll be added.", "error")
        return redirect(url_for("tax_invoice_print_options", invoice_id=invoice_id))

    # Bottom-left of the invoice prints the selected RTO's 4-line address
    # (Setup > 5. RTO Master) instead of the company address — but never
    # the RTO's own name/label, only its address lines.
    rto_address = None
    if ti.rto_name:
        rto = SimpleMaster.query.filter_by(kind="rto", name=ti.rto_name).first()
        if rto and rto.address:
            rto_address = rto.address

    # Bank details print from Bank Master (Setup > B) — the invoice's own
    # saved bank first, then whichever Bank is marked Default / Primary
    # (covers older invoices saved before this was wired up).
    bank_name = ti.bank_name
    bank_account_no = ti.bank_account_no
    bank_ifsc = ti.bank_ifsc
    if not bank_name:
        default_bank = SimpleMaster.query.filter_by(kind="bank", is_default=True).first()
        if default_bank:
            bank_name = default_bank.name
            bank_account_no = default_bank.account_no
            bank_ifsc = default_bank.ifsc

    # Model logo (Setup > 3. Product Master + static/logos/) — same
    # fetch-logo lookup already used on the Delivery Challan.
    product = Product.query.filter_by(name=ti.product_name).first()
    model_logo_url = _find_model_logo(product)

    return render_template("vouchers/tax_invoice_print.html", ti=ti, company=company,
                            doc_title=meta["title"], doc_no_label=meta["no_label"],
                            rto_address=rto_address, print_bank_name=bank_name,
                            print_bank_account_no=bank_account_no, print_bank_ifsc=bank_ifsc,
                            model_logo_url=model_logo_url)


@app.route("/m/tax-invoice/<int:invoice_id>/affidavit", endpoint="tax_invoice_affidavit")
def tax_invoice_affidavit(invoice_id):
    ti = TaxInvoice.query.get_or_404(invoice_id)
    company = Company.query.first()
    return render_template("vouchers/affidavit.html", ti=ti, company=company)


@app.route("/m/tax-invoice/<int:invoice_id>/undertaking", endpoint="tax_invoice_undertaking")
def tax_invoice_undertaking(invoice_id):
    ti = TaxInvoice.query.get_or_404(invoice_id)
    company = Company.query.first()
    return render_template("vouchers/undertaking.html", ti=ti, company=company)


@app.route("/m/tax-invoice/<int:invoice_id>/form22", endpoint="tax_invoice_form22")
def tax_invoice_form22(invoice_id):
    ti = TaxInvoice.query.get_or_404(invoice_id)
    company = Company.query.first()
    product = Product.query.filter_by(name=ti.product_name).first()
    return render_template("vouchers/form22.html", ti=ti, company=company, product=product)


@app.route("/m/tax-invoice/<int:invoice_id>/upload-code", endpoint="tax_invoice_upload_code")
def tax_invoice_upload_code(invoice_id):
    """Mirrors the original software's Invoice List 'Code ... is created for
    upload in File' popup: builds the pipe-delimited UMRN upload code and
    hands it back as a downloadable .txt, named after the Bill No. (the
    original saved it as '<BillNo>.TXT', e.g. 'S &0399.TXT').

    Format: UMRN|Chassis No.|Motor No.|MM/YYYY|R1|Colour Code|NA
      - UMRN comes from the Product Master's UMRN Code field (product.umrn_code)
      - MM/YYYY is the invoice date's month/year
      - "R1" is a fixed literal segment
      - Colour Code comes from the vehicle/invoice colour
      - "NA" is always literally "NA"
    """
    from flask import Response
    ti = TaxInvoice.query.get_or_404(invoice_id)

    product = Product.query.filter_by(name=ti.product_name).first()
    umrn = (product.umrn_code if product else None) or ""
    colour_code = (ti.vehicle.colour_code if ti.vehicle else None) or ""
    month_year = ti.date.strftime("%m/%Y") if ti.date else ""

    code = f"{umrn}|{ti.chassis_no or ''}|{ti.motor_no or ''}|{month_year}|R1|{colour_code}|NA"

    if not umrn:
        flash("Note: this Product has no UMRN Code set (Setup > Product Master) — that segment was left blank.", "error")

    filename = f"{(ti.bill_no or 'upload').replace('/', '_')}.TXT"
    return Response(
        code, mimetype="text/plain",
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""},
    )


# ---------------------------------------------------------------------------
# Purchase Bills — Vouchers > C
# Recording a raw-material purchase from a party. Independent of the
# vehicle pipeline — feeds the (future) Purchase Register report.
# ---------------------------------------------------------------------------
@app.route("/m/purchase-bills/full", methods=["GET"], endpoint="custom_purchase_bills")
def purchase_bill_list():
    bills = PurchaseBill.query.order_by(PurchaseBill.date.desc(), PurchaseBill.id.desc()).all()
    parties = SimpleMaster.query.filter_by(kind="party").order_by(SimpleMaster.name).all()
    raw_materials = Product.query.filter(
        (Product.fro == "R") | (Product.fro.is_(None))
    ).order_by(Product.name).all()
    next_no = (db.session.query(db.func.max(PurchaseBill.id)).scalar() or 0) + 1
    suggested_bill_no = f"PB-{next_no + 1000}"
    bills_by_id = {
        b.id: {
            "bill_no": b.bill_no or "",
            "date": b.date.isoformat() if b.date else "",
            "party_name": b.party_name or "",
            "party_gst_no": b.party_gst_no or "",
            "party_state_code": b.party_state_code or "07",
            "remarks": b.remarks or "",
            "items": [
                {"item_name": it.item_name, "hsn_code": it.hsn_code or "",
                 "qty": it.qty or 1, "rate": it.rate or 0, "gst_rate": it.gst_rate or 0}
                for it in b.items
            ],
        }
        for b in bills
    }
    return render_template(
        "vouchers/purchase_bill.html",
        bills=bills, parties=parties, raw_materials=raw_materials,
        suggested_bill_no=suggested_bill_no, today_iso=date.today().isoformat(),
        bills_by_id=bills_by_id,
    )


@app.route("/m/purchase-bills/new", methods=["POST"])
def purchase_bill_create():
    party_name = request.form.get("party_name", "").strip()
    if not party_name:
        flash("Party Name is required.", "error")
        return redirect(url_for("custom_purchase_bills"))

    item_names = request.form.getlist("item_name[]")
    hsn_codes = request.form.getlist("hsn_code[]")
    qtys = request.form.getlist("qty[]")
    rates = request.form.getlist("rate[]")
    gst_rates = request.form.getlist("gst_rate[]")

    pb = PurchaseBill(
        bill_no=request.form.get("bill_no"),
        date=_parse_date(request.form.get("date")),
        party_name=party_name,
        party_gst_no=request.form.get("party_gst_no"),
        party_state_code=request.form.get("party_state_code") or "07",
        remarks=request.form.get("remarks"),
    )

    added_any = False
    for i in range(len(item_names)):
        name = item_names[i].strip()
        if not name:
            continue
        pb.items.append(PurchaseBillItem(
            item_name=name,
            hsn_code=hsn_codes[i] if i < len(hsn_codes) else None,
            qty=float(qtys[i] or 1) if i < len(qtys) and qtys[i] else 1,
            rate=float(rates[i] or 0) if i < len(rates) and rates[i] else 0,
            gst_rate=float(gst_rates[i] or 0) if i < len(gst_rates) and gst_rates[i] else 0,
        ))
        added_any = True

    if not added_any:
        flash("Add at least one item line.", "error")
        return redirect(url_for("custom_purchase_bills"))

    db.session.add(pb)
    db.session.commit()
    flash(f"Purchase Bill {pb.bill_no or pb.id} saved with {len(pb.items)} item(s).", "success")
    return redirect(url_for("custom_purchase_bills"))


@app.route("/m/purchase-bills/<int:bill_id>/update", methods=["POST"], endpoint="purchase_bill_update")
def purchase_bill_update(bill_id):
    """Modify an existing Purchase Bill — replaces its header fields and item
    lines wholesale with whatever was submitted (simplest safe approach for
    an editable list of lines)."""
    pb = PurchaseBill.query.get_or_404(bill_id)

    party_name = request.form.get("party_name", "").strip()
    if not party_name:
        flash("Party Name is required.", "error")
        return redirect(url_for("custom_purchase_bills"))

    item_names = request.form.getlist("item_name[]")
    hsn_codes = request.form.getlist("hsn_code[]")
    qtys = request.form.getlist("qty[]")
    rates = request.form.getlist("rate[]")
    gst_rates = request.form.getlist("gst_rate[]")

    pb.bill_no = request.form.get("bill_no")
    pb.date = _parse_date(request.form.get("date")) or pb.date
    pb.party_name = party_name
    pb.party_gst_no = request.form.get("party_gst_no")
    pb.party_state_code = request.form.get("party_state_code") or "07"
    pb.remarks = request.form.get("remarks")

    # Replace item lines wholesale
    pb.items.clear()
    added_any = False
    for i in range(len(item_names)):
        name = item_names[i].strip()
        if not name:
            continue
        pb.items.append(PurchaseBillItem(
            item_name=name,
            hsn_code=hsn_codes[i] if i < len(hsn_codes) else None,
            qty=float(qtys[i] or 1) if i < len(qtys) and qtys[i] else 1,
            rate=float(rates[i] or 0) if i < len(rates) and rates[i] else 0,
            gst_rate=float(gst_rates[i] or 0) if i < len(gst_rates) and gst_rates[i] else 0,
        ))
        added_any = True

    if not added_any:
        flash("Add at least one item line.", "error")
        return redirect(url_for("custom_purchase_bills"))

    db.session.commit()
    flash(f"Purchase Bill {pb.bill_no or pb.id} updated.", "success")
    return redirect(url_for("custom_purchase_bills"))


@app.route("/m/purchase-bills/<int:bill_id>/delete", methods=["POST"])
def purchase_bill_delete(bill_id):
    pb = PurchaseBill.query.get_or_404(bill_id)
    db.session.delete(pb)
    db.session.commit()
    flash("Purchase Bill deleted.", "success")
    return redirect(url_for("custom_purchase_bills"))


# ---------------------------------------------------------------------------
# Old Rickshaw — Vouchers > G
# Buy-back / trade-in of a used e-rickshaw, tracked by vehicle registration
# number (not chassis) — independent of the manufacturing pipeline.
# ---------------------------------------------------------------------------
@app.route("/m/old-rickshaw/full", methods=["GET"], endpoint="custom_old_rickshaw")
def old_rickshaw_list():
    records = OldRickshaw.query.order_by(OldRickshaw.date.desc(), OldRickshaw.id.desc()).all()
    parties = SimpleMaster.query.filter_by(kind="party").order_by(SimpleMaster.name).all()
    next_no = (db.session.query(db.func.max(OldRickshaw.id)).scalar() or 0) + 1
    suggested_vou_no = str(next_no + 90)
    return render_template("vouchers/old_rickshaw.html", records=records, parties=parties,
                            suggested_vou_no=suggested_vou_no, today_iso=date.today().isoformat())


@app.route("/m/old-rickshaw/new", methods=["POST"])
def old_rickshaw_create():
    vehicle_reg_no = request.form.get("vehicle_reg_no", "").strip()
    if not vehicle_reg_no:
        flash("Vehicle Reg. No. is required.", "error")
        return redirect(url_for("custom_old_rickshaw"))

    rec = OldRickshaw(
        vou_no=request.form.get("vou_no"),
        date=_parse_date(request.form.get("date")),
        party_name=request.form.get("party_name"),
        vehicle_reg_no=vehicle_reg_no,
        model_name=request.form.get("model_name"),
        owner_name=request.form.get("owner_name"),
        salesman=request.form.get("salesman"),
        sold_amount=float(request.form.get("sold_amount") or 0),
        loan_amount=float(request.form.get("loan_amount") or 0),
        receipt_amount=float(request.form.get("receipt_amount") or 0),
        receipt_no=request.form.get("receipt_no"),
        ledger=request.form.get("ledger"),
        resale_date=_parse_date(request.form.get("resale_date")) if request.form.get("resale_date") else None,
        resale_ledger=request.form.get("resale_ledger"),
        remarks1=request.form.get("remarks1"),
        remarks2=request.form.get("remarks2"),
    )
    db.session.add(rec)
    db.session.commit()
    flash(f"Old Rickshaw record saved for {vehicle_reg_no}.", "success")
    return redirect(url_for("custom_old_rickshaw"))


@app.route("/m/old-rickshaw/<int:record_id>/delete", methods=["POST"])
def old_rickshaw_delete(record_id):
    rec = OldRickshaw.query.get_or_404(record_id)
    db.session.delete(rec)
    db.session.commit()
    flash("Old Rickshaw record deleted.", "success")
    return redirect(url_for("custom_old_rickshaw"))


# ---------------------------------------------------------------------------
# Battery Delivery Challan — Vouchers > H
# A battery sent to a dealer on its own (e.g. warranty replacement).
# ---------------------------------------------------------------------------
@app.route("/m/battery-delivery-challan/full", methods=["GET"], endpoint="custom_battery_delivery_challan")
def battery_delivery_challan_list():
    records = BatteryDeliveryChallan.query.order_by(
        BatteryDeliveryChallan.date.desc(), BatteryDeliveryChallan.id.desc()).all()
    dealers = Dealer.query.order_by(Dealer.name).all()
    battery_makers = SimpleMaster.query.filter_by(kind="battery-maker").order_by(SimpleMaster.name).all()
    next_no = (db.session.query(db.func.max(BatteryDeliveryChallan.id)).scalar() or 0) + 1
    suggested_challan_no = f"BDC{next_no + 1000}"
    return render_template("vouchers/battery_delivery_challan.html", records=records, dealers=dealers,
                            battery_makers=battery_makers, suggested_challan_no=suggested_challan_no,
                            today_iso=date.today().isoformat())


@app.route("/m/battery-delivery-challan/new", methods=["POST"])
def battery_delivery_challan_create():
    dealer_id = request.form.get("dealer_id", type=int)
    if not dealer_id:
        flash("Dealer is required.", "error")
        return redirect(url_for("custom_battery_delivery_challan"))

    rec = BatteryDeliveryChallan(
        challan_no=request.form.get("challan_no"),
        date=_parse_date(request.form.get("date")),
        dealer_id=dealer_id,
        battery_maker=request.form.get("battery_maker"),
        battery_no=request.form.get("battery_no"),
        qty=int(request.form.get("qty") or 1),
        remarks=request.form.get("remarks"),
    )
    db.session.add(rec)
    db.session.commit()
    flash(f"Battery Delivery Challan {rec.challan_no} saved.", "success")
    return redirect(url_for("custom_battery_delivery_challan"))


@app.route("/m/battery-delivery-challan/<int:record_id>/delete", methods=["POST"])
def battery_delivery_challan_delete(record_id):
    rec = BatteryDeliveryChallan.query.get_or_404(record_id)
    db.session.delete(rec)
    db.session.commit()
    flash("Battery Delivery Challan deleted.", "success")
    return redirect(url_for("custom_battery_delivery_challan"))


# ---------------------------------------------------------------------------
# Journal Stock — Vouchers > I
# Manual stock correction (raw material or finished item), for anything that
# doesn't come from a Purchase Bill, Production Voucher, or Delivery Challan.
# ---------------------------------------------------------------------------
@app.route("/m/journal-stock/full", methods=["GET"], endpoint="custom_journal_stock")
def journal_stock_list():
    records = JournalStock.query.order_by(JournalStock.date.desc(), JournalStock.id.desc()).all()
    all_items = Product.query.order_by(Product.name).all()
    next_no = (db.session.query(db.func.max(JournalStock.id)).scalar() or 0) + 1
    suggested_vou_no = f"J-{next_no + 100}"
    return render_template("vouchers/journal_stock.html", records=records, all_items=all_items,
                            suggested_vou_no=suggested_vou_no, today_iso=date.today().isoformat())


@app.route("/m/journal-stock/new", methods=["POST"])
def journal_stock_create():
    item_name = request.form.get("item_name", "").strip()
    if not item_name:
        flash("Item Name is required.", "error")
        return redirect(url_for("custom_journal_stock"))

    rec = JournalStock(
        vou_no=request.form.get("vou_no"),
        date=_parse_date(request.form.get("date")),
        item_name=item_name,
        item_type=request.form.get("item_type") or "R",
        qty=float(request.form.get("qty") or 0),
        reason=request.form.get("reason"),
    )
    db.session.add(rec)
    db.session.commit()
    flash(f"Journal Stock entry saved for {item_name}.", "success")
    return redirect(url_for("custom_journal_stock"))


@app.route("/m/journal-stock/<int:record_id>/delete", methods=["POST"])
def journal_stock_delete(record_id):
    rec = JournalStock.query.get_or_404(record_id)
    db.session.delete(rec)
    db.session.commit()
    flash("Journal Stock entry deleted.", "success")
    return redirect(url_for("custom_journal_stock"))


# ---------------------------------------------------------------------------
# Stock — J/K/L: Closing Stock (read-only, computed from existing vouchers)
# ---------------------------------------------------------------------------
@app.route("/m/closing-stock-premises/full", endpoint="custom_closing_stock_premises")
def closing_stock_premises():
    """Finished vehicles still at the factory — never left on a Delivery Challan."""
    vehicles = Vehicle.query.filter_by(stage="Manufacturing").order_by(Vehicle.model_name, Vehicle.colour).all()
    summary = {}
    for v in vehicles:
        key = (v.model_name or "—", v.colour or "—")
        summary[key] = summary.get(key, 0) + 1
    return render_template("stock/closing_stock_premises.html", vehicles=vehicles, summary=summary)


@app.route("/m/closing-stock-dealers/full", endpoint="custom_closing_stock_dealers")
def closing_stock_dealers():
    """Finished vehicles out on Delivery Challan but not yet invoiced (still dealer's stock)."""
    vehicles = Vehicle.query.filter_by(stage="Delivery Challan").order_by(Vehicle.dealer_name, Vehicle.model_name).all()
    summary = {}
    for v in vehicles:
        key = (v.dealer_name or "—", v.model_name or "—")
        summary[key] = summary.get(key, 0) + 1
    return render_template("stock/closing_stock_dealers.html", vehicles=vehicles, summary=summary)


@app.route("/m/closing-stock-raw/full", endpoint="custom_closing_stock_raw")
def closing_stock_raw():
    """Raw material on hand = total purchased - total consumed by Production Vouchers."""
    purchased = {}
    for item in PurchaseBillItem.query.all():
        key = item.item_name
        d = purchased.setdefault(key, {"hsn": item.hsn_code, "purchased": 0.0, "consumed": 0.0})
        d["purchased"] += item.qty or 0
        if item.hsn_code and not d["hsn"]:
            d["hsn"] = item.hsn_code

    for item in ProductionVoucherItem.query.all():
        key = item.item_name
        d = purchased.setdefault(key, {"hsn": None, "purchased": 0.0, "consumed": 0.0})
        d["consumed"] += item.qty or 0

    for j in JournalStock.query.filter_by(item_type="R").all():
        key = j.item_name
        d = purchased.setdefault(key, {"hsn": None, "purchased": 0.0, "consumed": 0.0})
        if j.qty >= 0:
            d["purchased"] += j.qty
        else:
            d["consumed"] += abs(j.qty)

    rows = []
    for name, d in sorted(purchased.items()):
        rows.append({
            "name": name, "hsn": d["hsn"], "purchased": d["purchased"],
            "consumed": d["consumed"], "closing": round(d["purchased"] - d["consumed"], 2),
        })
    return render_template("stock/closing_stock_raw.html", rows=rows)


# ---------------------------------------------------------------------------
# Stock — M/N: Stock Ledger (chronological IN/OUT movement, with running balance)
# ---------------------------------------------------------------------------
@app.route("/m/stock-ledger-premises/full", endpoint="custom_stock_ledger_premises")
def stock_ledger_premises():
    """
    Movement of finished vehicles in & out of the factory:
    IN  = Production Voucher (manufactured)
    OUT = Delivery Challan (left for a dealer)
    """
    from_date = _parse_date(request.args.get("from")) if request.args.get("from") else None
    to_date = _parse_date(request.args.get("to")) if request.args.get("to") else None

    events = []
    for pv in ProductionVoucher.query.all():
        if from_date and pv.date and pv.date < from_date:
            continue
        if to_date and pv.date and pv.date > to_date:
            continue
        events.append({"date": pv.date, "type": "IN", "doc_no": pv.vou_no,
                       "chassis_no": pv.chassis_no, "particulars": f"Production — {pv.product_name}", "qty": pv.quantity or 1})
    for dc in DeliveryChallan.query.filter_by(cancelled=False).all():
        if from_date and dc.date and dc.date < from_date:
            continue
        if to_date and dc.date and dc.date > to_date:
            continue
        events.append({"date": dc.date, "type": "OUT", "doc_no": dc.challan_no,
                       "chassis_no": dc.chassis_no,
                       "particulars": f"Delivery Challan to {dc.dealer.name if dc.dealer else ''}", "qty": 1})

    events.sort(key=lambda e: (e["date"] or date.min))
    balance = 0
    for e in events:
        balance += e["qty"] if e["type"] == "IN" else -e["qty"]
        e["balance"] = balance

    return render_template("stock/stock_ledger.html", title="Stock Ledger - Premises",
                            events=events, from_date=request.args.get("from", ""), to_date=request.args.get("to", ""),
                            filter_action=url_for("custom_stock_ledger_premises"))


@app.route("/m/stock-ledger-dealers/full", endpoint="custom_stock_ledger_dealers")
def stock_ledger_dealers():
    """
    Movement of finished vehicles in & out of a dealer's stock:
    IN  = Delivery Challan (arrived at dealer)
    OUT = Tax Invoice (sold, left dealer stock)
    """
    dealer_id = request.args.get("dealer_id", type=int)
    from_date = _parse_date(request.args.get("from")) if request.args.get("from") else None
    to_date = _parse_date(request.args.get("to")) if request.args.get("to") else None
    dealers = Dealer.query.order_by(Dealer.name).all()

    events = []
    dc_query = DeliveryChallan.query.filter_by(cancelled=False)
    if dealer_id:
        dc_query = dc_query.filter_by(dealer_id=dealer_id)
    for dc in dc_query.all():
        if from_date and dc.date and dc.date < from_date:
            continue
        if to_date and dc.date and dc.date > to_date:
            continue
        events.append({"date": dc.date, "type": "IN", "doc_no": dc.challan_no,
                       "chassis_no": dc.chassis_no, "dealer_name": dc.dealer.name if dc.dealer else "",
                       "particulars": f"Delivery Challan — {dc.product_name}", "qty": 1})

    ti_query = TaxInvoice.query.filter_by(cancelled=False)
    for ti in ti_query.all():
        if dealer_id and (not ti.delivery_challan or ti.delivery_challan.dealer_id != dealer_id):
            continue
        if from_date and ti.date and ti.date < from_date:
            continue
        if to_date and ti.date and ti.date > to_date:
            continue
        events.append({"date": ti.date, "type": "OUT", "doc_no": ti.bill_no,
                       "chassis_no": ti.chassis_no, "dealer_name": ti.dealer_name or "",
                       "particulars": f"Tax Invoice — {ti.product_name}", "qty": 1})

    events.sort(key=lambda e: (e["date"] or date.min))
    running = {}
    for e in events:
        key = e.get("dealer_name", "")
        running[key] = running.get(key, 0) + (e["qty"] if e["type"] == "IN" else -e["qty"])
        e["balance"] = running[key]

    return render_template("stock/stock_ledger_dealers.html",
                            events=events, dealers=dealers, selected_dealer_id=dealer_id,
                            from_date=request.args.get("from", ""), to_date=request.args.get("to", ""))


# ---------------------------------------------------------------------------
# Reports — O through X. All read-only views computed from the vouchers
# already entered (Purchase Bills, Production Vouchers, Delivery Challans,
# Tax Invoices), except Password which is a small admin utility.
# ---------------------------------------------------------------------------
def _date_filter_bounds():
    from_date = _parse_date(request.args.get("from")) if request.args.get("from") else None
    to_date = _parse_date(request.args.get("to")) if request.args.get("to") else None
    return from_date, to_date


def _in_range(d, from_date, to_date):
    if from_date and d and d < from_date:
        return False
    if to_date and d and d > to_date:
        return False
    return True


def _matches(search, *fields):
    """Case-insensitive substring match of `search` against any of `fields`."""
    if not search:
        return True
    needle = search.strip().lower()
    return any(needle in str(f).lower() for f in fields if f)


@app.route("/m/purchase-register/full", endpoint="custom_purchase_register")
def purchase_register():
    from_date, to_date = _date_filter_bounds()
    search = request.args.get("search", "").strip()
    rows = []
    for b in PurchaseBill.query.order_by(PurchaseBill.date).all():
        if not _in_range(b.date, from_date, to_date):
            continue
        if not _matches(search, b.party_name, b.bill_no):
            continue
        for it in b.items:
            rows.append({"date": b.date, "bill_no": b.bill_no or ".", "party_name": b.party_name,
                         "item_name": it.item_name, "hsn": it.hsn_code, "taxable_amt": it.taxable_amt,
                         "gst_rate": it.gst_rate, "is_inter_state": it.is_inter_state,
                         "cgst_amt": it.cgst_amt, "sgst_amt": it.sgst_amt, "igst_amt": it.igst_amt})
    totals = {
        "taxable": round(sum(r["taxable_amt"] for r in rows), 2),
        "cgst": round(sum(r["cgst_amt"] for r in rows), 2),
        "sgst": round(sum(r["sgst_amt"] for r in rows), 2),
        "igst": round(sum(r["igst_amt"] for r in rows), 2),
    }
    return render_template("reports/purchase_register.html", rows=rows, totals=totals,
                            from_date=request.args.get("from", ""), to_date=request.args.get("to", ""),
                            search=search,
                            filter_action=url_for("custom_purchase_register"),
                            export_action=url_for("purchase_register_export"))


@app.route("/m/purchase-register/export", endpoint="purchase_register_export")
def purchase_register_export():
    from_date, to_date = _date_filter_bounds()
    search = request.args.get("search", "").strip()
    rows = []
    for b in PurchaseBill.query.order_by(PurchaseBill.date).all():
        if not _in_range(b.date, from_date, to_date):
            continue
        if not _matches(search, b.party_name, b.bill_no):
            continue
        for it in b.items:
            rows.append([b.date.strftime("%d/%m/%Y") if b.date else "", b.bill_no or ".", b.party_name,
                         it.item_name, it.hsn_code or "", it.taxable_amt, it.cgst_amt, it.sgst_amt, it.igst_amt])
    headers = ["Date", "Bill No.", "Party Name", "Item Name", "HSN", "Taxable Amt", "CGST Amt", "SGST Amt", "IGST Amt"]
    return _export_excel("Purchase_Register.xlsx", headers, rows)


@app.route("/m/production-register/full", endpoint="custom_production_register")
def production_register():
    from_date, to_date = _date_filter_bounds()
    search = request.args.get("search", "").strip()
    vouchers = [v for v in ProductionVoucher.query.order_by(ProductionVoucher.date).all()
                if _in_range(v.date, from_date, to_date) and _matches(search, v.product_name, v.chassis_no)]
    return render_template("reports/production_register.html", vouchers=vouchers,
                            from_date=request.args.get("from", ""), to_date=request.args.get("to", ""),
                            search=search,
                            filter_action=url_for("custom_production_register"),
                            export_action=url_for("production_register_export"))


@app.route("/m/production-register/export", endpoint="production_register_export")
def production_register_export():
    from_date, to_date = _date_filter_bounds()
    search = request.args.get("search", "").strip()
    vouchers = [v for v in ProductionVoucher.query.order_by(ProductionVoucher.date).all()
                if _in_range(v.date, from_date, to_date) and _matches(search, v.product_name, v.chassis_no)]
    headers = ["Date", "Vou. No.", "Product Name", "Quantity", "Chassis No.", "Motor No.", "Controller No."]
    rows = [[v.date.strftime("%d/%m/%Y") if v.date else "", v.vou_no, v.product_name, v.quantity,
             v.chassis_no, v.motor_no, v.controller_no] for v in vouchers]
    return _export_excel("Production_Register.xlsx", headers, rows)


@app.route("/m/delivery-challan-register/full", endpoint="custom_delivery_challan_register")
def delivery_challan_register():
    from_date, to_date = _date_filter_bounds()
    search = request.args.get("search", "").strip()
    challans = [c for c in DeliveryChallan.query.order_by(DeliveryChallan.date).all()
                if _in_range(c.date, from_date, to_date)
                and _matches(search, c.dealer.name if c.dealer else None, c.chassis_no, c.challan_no)]
    # Bill No. (if billed) — prefer the actual Tax Invoice raised against this challan;
    # fall back to the challan's own sale_bill_no snapshot if no invoice exists yet.
    invoiced_bill_no = {ti.delivery_challan_id: ti.bill_no
                         for ti in TaxInvoice.query.filter(TaxInvoice.delivery_challan_id.isnot(None)).all()}
    bill_no_by_challan = {c.id: (invoiced_bill_no.get(c.id) or c.sale_bill_no or None) for c in challans}

    status = request.args.get("status", "all")
    if status == "sold":
        challans = [c for c in challans if bill_no_by_challan.get(c.id)]
    elif status == "unsold":
        challans = [c for c in challans if not bill_no_by_challan.get(c.id)]

    return render_template("reports/delivery_challan_register.html", challans=challans,
                            invoiced_bill_no=invoiced_bill_no, bill_no_by_challan=bill_no_by_challan,
                            status=status, search=search,
                            from_date=request.args.get("from", ""), to_date=request.args.get("to", ""),
                            filter_action=url_for("custom_delivery_challan_register"),
                            export_action=url_for("delivery_challan_register_export"))


@app.route("/m/delivery-challan-register/export", endpoint="delivery_challan_register_export")
def delivery_challan_register_export():
    from_date, to_date = _date_filter_bounds()
    search = request.args.get("search", "").strip()
    challans = [c for c in DeliveryChallan.query.order_by(DeliveryChallan.date).all()
                if _in_range(c.date, from_date, to_date)
                and _matches(search, c.dealer.name if c.dealer else None, c.chassis_no, c.challan_no)]
    invoiced_bill_no = {ti.delivery_challan_id: ti.bill_no
                         for ti in TaxInvoice.query.filter(TaxInvoice.delivery_challan_id.isnot(None)).all()}
    headers = ["Date", "Challan No.", "Dealer", "Product Name", "Chassis No.", "Bill No. (if billed)"]
    rows = [[c.date.strftime("%d/%m/%Y") if c.date else "", c.challan_no,
             c.dealer.name if c.dealer else "", c.product_name, c.chassis_no,
             invoiced_bill_no.get(c.id, "")] for c in challans]
    return _export_excel("Delivery_Challan_Register.xlsx", headers, rows)


@app.route("/m/sale-register/full", endpoint="custom_sale_register")
def sale_register():
    from_date, to_date = _date_filter_bounds()
    search = request.args.get("search", "").strip()
    invoices = [i for i in TaxInvoice.query.order_by(TaxInvoice.date).all()
                if _in_range(i.date, from_date, to_date)
                and _matches(search, i.buyer_name, i.dealer_name, i.product_name, i.chassis_no, i.bill_no)]
    totals = {
        "taxable": round(sum(i.taxable_value for i in invoices), 2),
        "tax": round(sum(i.tax_amount for i in invoices), 2),
        "insurance": round(sum(i.insurance_amount or 0 for i in invoices), 2),
        "registration": round(sum(i.registration_amount or 0 for i in invoices), 2),
        "total": round(sum(i.bill_total for i in invoices), 2),
    }
    return render_template("reports/sale_register.html", invoices=invoices, totals=totals,
                            from_date=request.args.get("from", ""), to_date=request.args.get("to", ""),
                            search=search,
                            filter_action=url_for("custom_sale_register"),
                            export_action=url_for("sale_register_export"))


@app.route("/m/sale-register/export", endpoint="sale_register_export")
def sale_register_export():
    from_date, to_date = _date_filter_bounds()
    search = request.args.get("search", "").strip()
    invoices = [i for i in TaxInvoice.query.order_by(TaxInvoice.date).all()
                if _in_range(i.date, from_date, to_date)
                and _matches(search, i.buyer_name, i.dealer_name, i.product_name, i.chassis_no, i.bill_no)]
    headers = ["Date", "Bill No.", "Buyer Name", "Product Name", "Chassis No.", "Taxable Value",
               "Tax Amount", "Insurance", "Registration", "Bill Total"]
    rows = [[i.date.strftime("%d/%m/%Y") if i.date else "", i.bill_no, i.buyer_name, i.product_name,
             i.chassis_no, i.taxable_value, i.tax_amount, i.insurance_amount or 0,
             i.registration_amount or 0, i.bill_total] for i in invoices]
    return _export_excel("Sale_Register.xlsx", headers, rows)


@app.route("/m/gst-register/full", endpoint="custom_gst_register")
def gst_register():
    """Outward supplies (Tax Invoices) vs Inward supplies (Purchase Bills) — for GST filing."""
    from_date, to_date = _date_filter_bounds()
    search = request.args.get("search", "").strip()
    outward = [i for i in TaxInvoice.query.filter_by(cancelled=False).order_by(TaxInvoice.date).all()
               if _in_range(i.date, from_date, to_date) and _matches(search, i.buyer_name, i.bill_no)]
    inward = []
    for b in PurchaseBill.query.order_by(PurchaseBill.date).all():
        if not _in_range(b.date, from_date, to_date):
            continue
        if not _matches(search, b.party_name, b.bill_no):
            continue
        for it in b.items:
            inward.append({"date": b.date, "doc_no": b.bill_no or ".", "party_name": b.party_name,
                           "taxable": it.taxable_amt, "cgst": it.cgst_amt, "sgst": it.sgst_amt, "igst": it.igst_amt})

    outward_totals = {
        "taxable": round(sum(i.taxable_value for i in outward), 2),
        "cgst": round(sum(i.cgst_amount for i in outward), 2),
        "sgst": round(sum(i.sgst_amount for i in outward), 2),
        "igst": round(sum(i.igst_amount for i in outward), 2),
    }
    inward_totals = {
        "taxable": round(sum(r["taxable"] for r in inward), 2),
        "cgst": round(sum(r["cgst"] for r in inward), 2),
        "sgst": round(sum(r["sgst"] for r in inward), 2),
        "igst": round(sum(r["igst"] for r in inward), 2),
    }
    return render_template("reports/gst_register.html", outward=outward, inward=inward,
                            outward_totals=outward_totals, inward_totals=inward_totals,
                            from_date=request.args.get("from", ""), to_date=request.args.get("to", ""),
                            search=search,
                            filter_action=url_for("custom_gst_register"),
                            export_action=url_for("gst_register_export"))


@app.route("/m/gst-register/export", endpoint="gst_register_export")
def gst_register_export():
    from io import BytesIO
    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.styles import Font

    from_date, to_date = _date_filter_bounds()
    search = request.args.get("search", "").strip()
    outward = [i for i in TaxInvoice.query.filter_by(cancelled=False).order_by(TaxInvoice.date).all()
               if _in_range(i.date, from_date, to_date) and _matches(search, i.buyer_name, i.bill_no)]
    inward = []
    for b in PurchaseBill.query.order_by(PurchaseBill.date).all():
        if not _in_range(b.date, from_date, to_date):
            continue
        if not _matches(search, b.party_name, b.bill_no):
            continue
        for it in b.items:
            inward.append({"date": b.date, "doc_no": b.bill_no or ".", "party_name": b.party_name,
                           "taxable": it.taxable_amt, "cgst": it.cgst_amt, "sgst": it.sgst_amt, "igst": it.igst_amt})

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Outward (Sales)"
    ws1.append(["Date", "Bill No.", "Party Name", "Taxable Value", "CGST", "SGST", "IGST"])
    for cell in ws1[1]:
        cell.font = Font(bold=True)
    for i in outward:
        ws1.append([i.date.strftime("%d/%m/%Y") if i.date else "", i.bill_no, i.buyer_name,
                    i.taxable_value, i.cgst_amount, i.sgst_amount, i.igst_amount])

    ws2 = wb.create_sheet("Inward (Purchases)")
    ws2.append(["Date", "Bill No.", "Party Name", "Taxable Value", "CGST", "SGST", "IGST"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for r in inward:
        ws2.append([r["date"].strftime("%d/%m/%Y") if r["date"] else "", r["doc_no"], r["party_name"],
                    r["taxable"], r["cgst"], r["sgst"], r["igst"]])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="GST_Register.xlsx",
                      mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/m/hypothecation-register/full", endpoint="custom_hypothecation_register")
def hypothecation_register():
    from_date, to_date = _date_filter_bounds()
    search = request.args.get("search", "").strip()
    invoices = [i for i in TaxInvoice.query.filter(TaxInvoice.financer_name.isnot(None),
                                                     TaxInvoice.financer_name != "").order_by(TaxInvoice.date).all()
                if _in_range(i.date, from_date, to_date)
                and _matches(search, i.buyer_name, i.financer_name, i.bill_no)]
    total_hyp = round(sum(i.hypothecation_amount or 0 for i in invoices), 2)
    return render_template("reports/hypothecation_register.html", invoices=invoices, total_hyp=total_hyp,
                            from_date=request.args.get("from", ""), to_date=request.args.get("to", ""),
                            search=search,
                            filter_action=url_for("custom_hypothecation_register"),
                            export_action=url_for("hypothecation_register_export"))


@app.route("/m/hypothecation-register/export", endpoint="hypothecation_register_export")
def hypothecation_register_export():
    from_date, to_date = _date_filter_bounds()
    search = request.args.get("search", "").strip()
    invoices = [i for i in TaxInvoice.query.filter(TaxInvoice.financer_name.isnot(None),
                                                     TaxInvoice.financer_name != "").order_by(TaxInvoice.date).all()
                if _in_range(i.date, from_date, to_date)
                and _matches(search, i.buyer_name, i.financer_name, i.bill_no)]
    headers = ["Date", "Bill No.", "Buyer Name", "Chassis No.", "Financer Name", "Hypothecation Amount"]
    rows = [[i.date.strftime("%d/%m/%Y") if i.date else "", i.bill_no, i.buyer_name, i.chassis_no,
             i.financer_name, i.hypothecation_amount or 0] for i in invoices]
    return _export_excel("Hypothecation_Register.xlsx", headers, rows)


@app.route("/m/payment-receivable-report/full", endpoint="custom_payment_receivable_report")
def payment_receivable_report():
    from_date, to_date = _date_filter_bounds()
    search = request.args.get("search", "").strip()
    show_all = request.args.get("show_all") == "1"
    invoices = [i for i in TaxInvoice.query.filter_by(cancelled=False).order_by(TaxInvoice.date).all()
                if _in_range(i.date, from_date, to_date)
                and _matches(search, i.dealer_name, i.buyer_name, i.bill_no)]
    if not show_all:
        invoices = [i for i in invoices if i.balance_due > 0]
    total_due = round(sum(i.balance_due for i in invoices), 2)
    total_value = round(sum(i.bill_total for i in invoices), 2)
    total_loan = round(sum(i.hypothecation_amount or 0 for i in invoices), 2)
    total_recd = round(sum(i.amount_received or 0 for i in invoices), 2)
    total_subsidy = round(sum(i.subsidy_amount or 0 for i in invoices), 2)

    # Snapshot for the "Payment Details" popup (click a row to open) — the
    # editable fields plus a few read-only display ones, keyed by invoice id.
    payments_by_id = {
        i.id: {
            "product_name": i.product_name or "", "chassis_no": i.chassis_no or "",
            "bill_no": i.bill_no or "",
            "sale_amount": i.sale_amount or 0, "hypothecation_amount": i.hypothecation_amount or 0,
            "amount_received": i.amount_received or 0, "subsidy_amount": i.subsidy_amount or 0,
            "subsidy_status": i.subsidy_status or "Due",
            "financer_name": i.financer_name or "", "voucher_no": i.voucher_no or "",
            "chassis_record_no": i.chassis_record_no or "", "ledger_no": i.ledger_no or "",
            "cancelled_cheque_no": i.cancelled_cheque_no or "", "vehicle_reg_no": i.vehicle_reg_no or "",
        }
        for i in invoices
    }

    return render_template("reports/payment_receivable_report.html", invoices=invoices, total_due=total_due,
                            total_value=total_value, total_loan=total_loan, total_recd=total_recd,
                            total_subsidy=total_subsidy,
                            show_all=show_all, search=search,
                            from_date=request.args.get("from", ""), to_date=request.args.get("to", ""),
                            export_action=url_for("payment_receivable_report_export"),
                            payments_by_id=payments_by_id)


@app.route("/m/payment-receivable-report/export", endpoint="payment_receivable_report_export")
def payment_receivable_report_export():
    from_date, to_date = _date_filter_bounds()
    search = request.args.get("search", "").strip()
    show_all = request.args.get("show_all") == "1"
    invoices = [i for i in TaxInvoice.query.filter_by(cancelled=False).order_by(TaxInvoice.date).all()
                if _in_range(i.date, from_date, to_date)
                and _matches(search, i.dealer_name, i.buyer_name, i.bill_no)]
    if not show_all:
        invoices = [i for i in invoices if i.balance_due > 0]
    headers = ["Date", "Bill No.", "Buyer Name", "Bill Total", "Loan Amount", "Amount Received",
               "Subsidy Amount", "Balance Due"]
    rows = [[i.date.strftime("%d/%m/%Y") if i.date else "", i.bill_no, i.buyer_name, i.bill_total,
             i.hypothecation_amount or 0, i.amount_received or 0, i.subsidy_amount or 0, i.balance_due]
            for i in invoices]
    return _export_excel("Payment_Receivable_Report.xlsx", headers, rows)


@app.route("/m/subsidy-report/full", endpoint="custom_subsidy_report")
def subsidy_report():
    from_date, to_date = _date_filter_bounds()
    search = request.args.get("search", "").strip()
    invoices = [i for i in TaxInvoice.query.filter(TaxInvoice.subsidy_amount > 0).order_by(TaxInvoice.date).all()
                if _in_range(i.date, from_date, to_date) and _matches(search, i.buyer_name, i.bill_no)]
    total_subsidy = round(sum(i.subsidy_amount or 0 for i in invoices), 2)
    return render_template("reports/subsidy_report.html", invoices=invoices, total_subsidy=total_subsidy,
                            from_date=request.args.get("from", ""), to_date=request.args.get("to", ""),
                            search=search,
                            filter_action=url_for("custom_subsidy_report"),
                            export_action=url_for("subsidy_report_export"))


@app.route("/m/subsidy-report/export", endpoint="subsidy_report_export")
def subsidy_report_export():
    from_date, to_date = _date_filter_bounds()
    search = request.args.get("search", "").strip()
    invoices = [i for i in TaxInvoice.query.filter(TaxInvoice.subsidy_amount > 0).order_by(TaxInvoice.date).all()
                if _in_range(i.date, from_date, to_date) and _matches(search, i.buyer_name, i.bill_no)]
    headers = ["Date", "Bill No.", "Buyer Name", "Chassis No.", "Subsidy Amount"]
    rows = [[i.date.strftime("%d/%m/%Y") if i.date else "", i.bill_no, i.buyer_name, i.chassis_no,
             i.subsidy_amount or 0] for i in invoices]
    return _export_excel("Subsidy_Report.xlsx", headers, rows)


@app.route("/m/ledger/full", endpoint="custom_ledger")
def ledger():
    """
    'W. Ledger' — mirrors the legacy desktop app's 'Stock Ledger' dialog,
    which offers 4 radio choices (Trial Balance / Account Statement /
    Receivable Report / Day Book Entry) and shows the chosen report/screen
    below once you hit Okay. Day Book Entry used to be its own separate
    menu item (W1.) in this web app; it's now reached the same way the
    desktop app does it -- as a nested popup underneath this chooser.
    """
    view = request.args.get("view", "account_statement")
    from_date, to_date = _date_filter_bounds()

    ctx = {
        "view": view,
        "from_date": request.args.get("from", ""),
        "to_date": request.args.get("to", ""),
    }

    # --- Account Statement (dealer ledger) ---------------------------------
    # Mirrors the real double-entry structure seen in the original software:
    # each Tax Invoice posts a DEBIT to "SALE" for the full bill amount; any
    # money received against it (financed via a Financer, or paid directly)
    # posts a separate CREDIT to "CASH". Dealers here run as advance/deposit
    # accounts, so a Credit balance ("Cr") is normal and INCREASES the
    # balance, while a Debit (a sale) DRAWS DOWN that balance — matching the
    # original: balance += credit - debit.
    dealer_id = request.args.get("dealer_id", type=int)
    dealers = Dealer.query.order_by(Dealer.name).all()
    selected_dealer = Dealer.query.get(dealer_id) if dealer_id else None

    events = []
    if dealer_id:
        query = TaxInvoice.query.join(DeliveryChallan, TaxInvoice.delivery_challan_id == DeliveryChallan.id) \
            .filter(DeliveryChallan.dealer_id == dealer_id, TaxInvoice.cancelled.is_(False))
        for ti in query.all():
            if not _in_range(ti.date, from_date, to_date):
                continue
            events.append({
                "date": ti.date, "vr_type": "S", "doc_no": f"BILL NO.{ti.bill_no}",
                "account": "SALE", "debit": ti.bill_total, "credit": 0,
                "lines": [ti.product_name, ti.chassis_no],
            })
            if ti.hypothecation_amount:
                events.append({
                    "date": ti.date, "vr_type": "S", "doc_no": f"BILL NO.{ti.bill_no}",
                    "account": "CASH", "debit": 0, "credit": ti.hypothecation_amount,
                    "lines": [ti.financer_name or "Financer"],
                })
            direct_received = (ti.amount_received or 0) - (ti.hypothecation_amount or 0)
            if direct_received > 0:
                ref = ti.cancelled_cheque_no or ""
                events.append({
                    "date": ti.date, "vr_type": "S", "doc_no": f"BILL NO.{ti.bill_no}",
                    "account": "CASH", "debit": 0, "credit": direct_received,
                    "lines": [ref] if ref else [],
                })

        if selected_dealer and selected_dealer.name:
            db_rows = DayBook.query.filter(
                db.func.lower(DayBook.dealer_name) == selected_dealer.name.strip().lower()).all()
            for db_row in db_rows:
                if not _in_range(db_row.date, from_date, to_date):
                    continue
                events.append({
                    "date": db_row.date, "vr_type": "C", "doc_no": "",
                    "account": "CASH", "debit": db_row.debit_paid or 0,
                    "credit": db_row.credit_received or 0,
                    "lines": [db_row.narration] if db_row.narration else [],
                })

    events.sort(key=lambda e: (e["date"] or date.min))
    balance = 0
    for e in events:
        balance += e["credit"] - e["debit"]
        e["balance"] = round(abs(balance), 2)
        e["dc"] = "Cr" if balance >= 0 else "Dr"

    ctx.update(events=events, dealers=dealers, selected_dealer_id=dealer_id)

    # --- Day Book Entry (nested popup list + form) --------------------------
    if view == "day_book":
        search = request.args.get("search", "").strip()
        dquery = DayBook.query
        if search:
            dquery = dquery.filter(DayBook.dealer_name.ilike(f"%{search}%"))
        day_book_entries = dquery.order_by(DayBook.date.desc(), DayBook.vr_no.desc()).all()
        ctx.update(
            day_book_entries=day_book_entries,
            day_book_search=search,
            dealer_names=[d.name for d in dealers],
            next_vr_no=DayBook.next_vr_no(),
            today=date.today().isoformat(),
        )

    return render_template("reports/ledger.html", **ctx)


@app.route("/m/day-book/full", methods=["GET", "POST"], endpoint="custom_day_book")
def day_book():
    if request.method == "POST":
        row_id = request.form.get("id")
        row = DayBook.query.get(int(row_id)) if row_id else DayBook()
        if not row_id:
            row.vr_no = DayBook.next_vr_no()
        date_str = request.form.get("date")
        row.date = dt.strptime(date_str, "%Y-%m-%d").date() if date_str else None
        row.dealer_name = request.form.get("dealer_name", "").strip()
        row.credit_received = float(request.form.get("credit_received") or 0)
        row.debit_paid = float(request.form.get("debit_paid") or 0)
        row.narration = request.form.get("narration") or None
        return_to = request.form.get("return_to") or url_for("custom_ledger", view="day_book")
        if not row.dealer_name:
            flash("Dealer Name is required.", "error")
            return redirect(return_to)
        db.session.add(row)
        db.session.commit()
        flash(f"Saved Day Book entry Vr.No. {row.vr_no}.", "success")
        return redirect(return_to)

    search = request.args.get("search", "").strip()
    query = DayBook.query
    if search:
        query = query.filter(DayBook.dealer_name.ilike(f"%{search}%"))
    entries = query.order_by(DayBook.date.desc(), DayBook.vr_no.desc()).all()
    dealer_names = [d.name for d in Dealer.query.order_by(Dealer.name).all()]
    next_vr_no = DayBook.next_vr_no()
    return render_template("vouchers/day_book.html", entries=entries, search=search,
                            dealer_names=dealer_names, next_vr_no=next_vr_no, today=date.today().isoformat())


@app.route("/m/day-book/delete/<int:row_id>", methods=["POST"], endpoint="day_book_delete")
def day_book_delete(row_id):
    row = DayBook.query.get_or_404(row_id)
    db.session.delete(row)
    db.session.commit()
    flash("Day Book entry deleted.", "success")
    return_to = request.form.get("return_to") or url_for("custom_ledger", view="day_book")
    return redirect(return_to)


@app.route("/m/password/full", methods=["GET", "POST"], endpoint="custom_password")
def password_utility():
    if request.method == "POST":
        user_id = request.form.get("user_id", type=int)
        new_password = request.form.get("new_password", "")
        confirm_password = request.form.get("confirm_password", "")
        u = User.query.get_or_404(user_id)
        if not new_password or new_password != confirm_password:
            flash("Passwords must match and not be blank.", "error")
            return redirect(url_for("custom_password"))
        u.set_password(new_password)
        db.session.commit()
        flash(f"Password updated for '{u.username}'.", "success")
        return redirect(url_for("custom_password"))

    users = User.query.order_by(User.username).all()
    return render_template("reports/password.html", users=users)


# ---------------------------------------------------------------------------
# Import Old Data — upload the legacy Excel export (and, later, the Access
# .mdb) and load it into the new database.
# ---------------------------------------------------------------------------
@app.route("/m/import-data/full", methods=["GET", "POST"], endpoint="custom_import_data")
def import_data():
    if request.method == "POST":
        f = request.files.get("data_file")
        if not f or f.filename == "":
            flash("Please choose a file first.", "error")
            return redirect(url_for("custom_import_data"))

        ext = os.path.splitext(f.filename)[1].lower()
        from import_excel import run_import, run_mdb_import, MdbAccessUnavailable

        try:
            if ext in (".mdb", ".accdb"):
                # Vercel's deployed filesystem is read-only. /tmp is writable
                # during a serverless invocation, so use it for temporary imports.
                path = os.path.join("/tmp", f"uploaded_import{ext}")
                f.save(path)
                mdb_password = request.form.get("mdb_password") or None
                summary = run_mdb_import(path, password=mdb_password)
            elif ext == ".xlsx":
                path = os.path.join("/tmp", "uploaded_import.xlsx")
                f.save(path)
                summary = run_import(path)
            else:
                flash("Unsupported file type — please upload .xlsx, .mdb, or .accdb.", "error")
                return redirect(url_for("custom_import_data"))
        except MdbAccessUnavailable as e:
            flash(str(e), "error")
            return redirect(url_for("custom_import_data"))

        flash("Import finished: " + ", ".join(f"{k} {v}" for k, v in summary.items()), "success")
        return redirect(url_for("custom_import_data"))

    counts = {
        "Dealers": Dealer.query.count(),
        "Party Master (Purchase Parties)": SimpleMaster.query.filter_by(kind="party").count(),
        "Products": Product.query.count(),
        "Vehicles": Vehicle.query.count(),
        "Users": User.query.count(),
        "Production Vouchers": ProductionVoucher.query.count(),
        "Production Formula": ProductionFormula.query.count(),
        "Delivery Challans": DeliveryChallan.query.count(),
        "Tax Invoices": TaxInvoice.query.count(),
        "Purchase Bills": PurchaseBill.query.count(),
        "Old Rickshaw Records": OldRickshaw.query.count(),
        "Battery Delivery Challans": BatteryDeliveryChallan.query.count(),
        "Journal Stock Entries": JournalStock.query.count(),
        "Day Book Entries": DayBook.query.count(),
    }
    return render_template("import_data.html", counts=counts)


def _auto_migrate():
    """
    SQLite doesn't update an existing table's schema when a model gains new
    columns. Rather than making you delete your real data every time this
    app is updated, this checks each table for columns the model expects
    but the database file doesn't have yet, and adds them with ALTER TABLE
    ADD COLUMN — safe, additive, and keeps all your existing rows intact.

    This only runs against SQLite. For PostgreSQL/MySQL (production setups),
    schema changes should go through a proper migration tool (e.g. Alembic /
    Flask-Migrate) rather than an automatic best-effort ALTER TABLE — the
    quoting/type rules below are SQLite-specific and aren't safe to run
    unmodified against those engines.
    """
    if db.engine.dialect.name != "sqlite":
        print(f"[auto-migrate] Skipped — using {db.engine.dialect.name}, not SQLite. "
              f"Add new columns via a proper migration (e.g. Alembic) instead.")
        return

    from sqlalchemy import inspect, text

    inspector = inspect(db.engine)
    existing_tables = set(inspector.get_table_names())

    # Map SQLAlchemy column types to SQLite column types for ALTER TABLE
    def sqlite_type(col):
        t = str(col.type)
        if "BOOLEAN" in t:
            return "BOOLEAN"
        if "INTEGER" in t:
            return "INTEGER"
        if "FLOAT" in t or "NUMERIC" in t:
            return "FLOAT"
        if "DATE" in t or "TIME" in t:
            return t  # DATE / DATETIME
        return "TEXT"

    added = []
    for table in db.metadata.sorted_tables:
        if table.name not in existing_tables:
            continue  # brand-new table — db.create_all() already handles this
        existing_cols = {c["name"] for c in inspector.get_columns(table.name)}
        for col in table.columns:
            if col.name in existing_cols:
                continue
            col_type = sqlite_type(col)
            default_clause = ""
            if col.default is not None and getattr(col.default, "arg", None) is not None \
                    and not callable(col.default.arg):
                val = col.default.arg
                default_clause = f" DEFAULT {val!r}" if isinstance(val, str) else f" DEFAULT {val}"
            ddl = f'ALTER TABLE "{table.name}" ADD COLUMN "{col.name}" {col_type}{default_clause}'
            db.session.execute(text(ddl))
            added.append(f"{table.name}.{col.name}")
    if added:
        db.session.commit()
        print(f"[auto-migrate] Added missing columns: {', '.join(added)}")

    # One-time backfill: rows created before formula_name existed get their
    # product_name as a sensible default (matches the legacy app, where every
    # pre-existing formula's name was the same as its finished product).
    if "production_formula.formula_name" in added:
        db.session.execute(text(
            'UPDATE production_formula SET formula_name = product_name '
            'WHERE formula_name IS NULL OR formula_name = ""'
        ))
        db.session.commit()
        print("[auto-migrate] Backfilled production_formula.formula_name from product_name")


def _seed_defaults():
    """Create the schema and safe first-run defaults."""
    db.create_all()
    _auto_migrate()
    if Company.query.first() is None:
        db.session.add(Company(name="G.R.D. MOTORS"))
        db.session.commit()
    if User.query.first() is None:
        # The deployment-safe default password can be overridden through an
        # environment variable. Change it immediately after first login.
        default_password = os.environ.get("DEFAULT_ADMIN_PASSWORD", "admin1")
        u = User(username="admin", is_super_user=True, permissions="Y Y")
        u.set_password(default_password)
        db.session.add(u)
        db.session.commit()


# Vercel imports this module as a serverless function. Initialize the schema
# once per warm instance so the very first request can use the tables.
try:
    with app.app_context():
        _seed_defaults()
except Exception as exc:
    # Do not hide the real error from Vercel; log it and let the request fail
    # with a useful database error if the environment is not configured yet.
    print(f"[startup] Database initialization skipped/failed: {exc}")


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
